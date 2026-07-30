from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime, timedelta
import json
import math
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import urllib.parse
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fare_store import load_fare_snapshot


def app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


BASE_DIR = app_base_dir()

def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    try:
        with temp_path.open("w", encoding="utf-8", newline="\n") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def resource_path(relative_path: str) -> Path:
    external_path = BASE_DIR / relative_path
    if external_path.exists():
        return external_path
    bundle_root = getattr(sys, "_MEIPASS", None)
    if bundle_root:
        bundled_path = Path(str(bundle_root)) / relative_path
        if bundled_path.exists():
            return bundled_path
    return external_path


def app_command_prefix() -> list[str]:
    if getattr(sys, "frozen", False):
        return [sys.executable]
    return [sys.executable, "-X", "utf8", str(Path(__file__).resolve())]


def collector_command_prefix() -> list[str]:
    if getattr(sys, "frozen", False):
        return [sys.executable, "--collector"]
    return [sys.executable, "-X", "utf8", str(BASE_DIR / "topas_live_collector.py")]


MANIFEST_FILENAME = "hotels-manifest.json"
HOTELS_MANIFEST_PATH = BASE_DIR / MANIFEST_FILENAME
MASTER_PATH = resource_path("flight-master.mjs")
OUTPUT_DIR = BASE_DIR / "output" / "excel"
LOGS_DIR = BASE_DIR / "output" / "logs"
RAW_DIR = BASE_DIR / "output" / "raw"
RAW_STORE_FILENAME = "raw-store.sqlite"
LOCAL_FARE_CACHE_PATH = BASE_DIR / "output" / "cache" / "fares_snapshot.json"
FARE_SEED_CACHE_PATH = resource_path("fares_snapshot.seed.json")
AUTO_COLLECT_BATCH_SIZE = 91
TOPAS_DEBUG_URL = "https://www.topassellconnect.com/"
TOPAS_DEBUG_PORT = 9222
TOPAS_DEBUG_ADDRESS = f"127.0.0.1:{TOPAS_DEBUG_PORT}"
TOPAS_DEBUG_FALLBACK_ADDRESSES = (TOPAS_DEBUG_ADDRESS,)
TOPAS_DEBUG_PROFILE_DIR = BASE_DIR / "output" / "chrome-debug-profile"
APP_ICON_ICO_PATH = resource_path("assets/air_auto_lookup_icon.ico")
APP_ICON_PNG_PATH = resource_path("assets/air_auto_lookup_icon.png")

MONTH_CODES = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
FARE_ROUTE_MAP = {
    "ICNDAD_LJ081/2": "ICN-DAD-LJ",
    "ICNDAD_7C2217/8": "ICN-DAD-7C",
    "ICNDAD_ZE593/4": "ICN-DAD-ZE",
    "ICNDAD_TW013/4": "ICN-DAD-TW",
    "ICNCXR_TW033/4": "ICN-CXR-TW",
    "ICNCXR_LJ087/8": "ICN-CXR-LJ",
    "ICNCXR_ZE561/2": "ICN-CXR-ZE",
    "ICNCXR_7C2303/4": "ICN-CXR-7C",
    "ICNPQC_LJ091/2": "ICN-PQC-LJ",
    "ICNPQC_7C2315/6": "ICN-PQC-7C",
    "ICNPQC_ZE581/2": "ICN-PQC-ZE",
    "ICNPQC_TW055/6": "ICN-PQC-TW",
    "ICNBKK_ZE511/2": "ICN-BKK-ZE",
    "ICNBKK_YP601/2": "ICN-BKK-YP",
    "ICNBKK_LJ001/2": "ICN-BKK-LJ",
    "ICNBKK_TW101/2": "ICN-BKK-TW",
    "ICNCNX_LJ027/8": "ICN-CNX-LJ",
    "ICNCNX_ZE517/8": "ICN-CNX-ZE",
    "ICNCNX_7C2515/6": "ICN-CNX-7C",
    "PUSDAD_LJ763/4": "PUS-DAD-LJ",
    "PUSCXR_TW041/2": "PUS-CXR-TW",
    "PUSBKK_LJ021/2": "PUS-BKK-LJ",
    "PUSBKK_7C2551/2": "PUS-BKK-7C",
    "PUSCNX_ZE917/8": "PUS-CNX-ZE",
}
TOPAS_FLIGHT_LINE_RE = re.compile(
    r"^\s*(?P<line_no>\d+)\s+"
    r"(?P<airline>[A-Z0-9]{2})\s*"
    r"(?P<flight>\d{1,4})\s+"
    r"(?P<class_text>.*?)\s+"
    r"(?:/\s*)?"
    r"(?P<origin>[A-Z]{3})\s+"
    r"(?P<origin_terminal>[A-Z0-9]+)?\s*"
    r"(?P<destination>[A-Z]{3})\s+"
    r"(?P<destination_terminal>[A-Z0-9]+)?\s+"
    r"(?P<depart_time>\d{4})\s+"
    r"(?P<arrive_time>\d{4})(?P<arrive_day_offset>[+-]\d+)?\s*"
    r"(?:(?P<meal>[A-Z0-9]+)\s*/\s*)?"
    r"(?P<equipment>[A-Z0-9.]{3,})\s+"
    r"(?:[A-Z]{2,4}\s+)*"
    r"(?P<duration>\d{1,2}:\d{2})",
    re.IGNORECASE,
)
TOPAS_FULL_COMMAND_RE = re.compile(
    r"AN(?P<date>\d{1,2}[A-Z]{3})(?P<origin>[A-Z]{3})(?P<destination>[A-Z]{3})"
    r"/A(?P<airline>[A-Z0-9]{2})(?P<flight>\d{0,4})",
    re.IGNORECASE,
)
TOPAS_CLASS_RE = re.compile(r"\b([A-Z])([0-9A-Z])\b", re.IGNORECASE)
RETRY_STATE_FILE = "retry-state.json"
DEFAULT_MAX_COLLECT_ERROR_RETRIES = 3
RAW_COMPLETE_MARKERS = ("AMADEUS AVAILABILITY", "NO FLIGHT", "REQUEST NEW AVAILABILITY")
PAUSE_CONTROL_FILE = "pause.flag"
STOP_CONTROL_FILE = "stop.flag"

@dataclass(frozen=True)
class FlightMaster:
    key: str
    raw_key: str
    origin: str
    destination: str
    route: str
    airline: str
    dep_flight: str
    ret_flight: str
    ret_departure_time: str | None
    default_product_days: int
    fare_route: str = ""
    enabled: bool = True


def load_flight_masters(path: Path | None = None) -> list[FlightMaster]:
    items = load_flight_master_items(path)
    masters = []
    for item in items:
        masters.append(
            FlightMaster(
                key=str(item["key"]),
                raw_key=str(item.get("rawKey", item["key"])),
                origin=str(item["origin"]).upper(),
                destination=str(item["destination"]).upper(),
                route=str(item["route"]).upper(),
                airline=str(item["airline"]).upper(),
                dep_flight=str(item["depFlight"]).upper(),
                ret_flight=str(item["retFlight"]).upper(),
                ret_departure_time=item.get("retDepartureTime"),
                default_product_days=int(item.get("defaultProductDays") or 5),
                fare_route=str(item.get("fareRoute") or "").upper(),
                enabled=bool(item.get("enabled", True)),
            )
        )
    return masters


def load_flight_master_items(path: Path | None = None) -> list[dict[str, object]]:
    source_path = path or flight_master_source_path()
    items = load_flight_master_items_from_path(source_path)
    return [normalize_flight_master_item(item) for item in items]


def flight_master_source_path() -> Path:
    if HOTELS_MANIFEST_PATH.exists():
        return HOTELS_MANIFEST_PATH
    bundled_manifest = resource_path(MANIFEST_FILENAME)
    if bundled_manifest.exists():
        return bundled_manifest
    return MASTER_PATH


def load_flight_master_items_from_path(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() == ".json":
        payload = json.loads(text)
        if isinstance(payload, list):
            return [dict(item) for item in payload]
        if isinstance(payload, dict):
            items = None
            for key in ("flightMasters", "routes", "masters"):
                if key in payload:
                    items = payload[key]
                    break
            if isinstance(items, list):
                return [dict(item) for item in items]
        raise ValueError(f"{path.name}에서 flightMasters 목록을 찾지 못했습니다.")

    array_text = _extract_js_array(text, "flightMasters")
    payload = _js_object_array_to_json(array_text)
    return [dict(item) for item in json.loads(payload)]


APP_VERSION = "v1.0.9"


DEFAULT_INITIAL_MASTERS = [
    {
        "key": "ICNGUM_LJ915_916",
        "rawKey": "ICNGUM_LJ915/6",
        "origin": "ICN",
        "destination": "GUM",
        "route": "ICN-GUM",
        "airline": "LJ",
        "depFlight": "LJ915",
        "retFlight": "LJ916",
        "retDepartureTime": None,
        "defaultProductDays": 5,
        "fareRoute": "ICN-GUM-LJ",
        "enabled": True
    },
    {
        "key": "ICNGUM_LJ915_918",
        "rawKey": "ICNGUM_LJ915/8",
        "origin": "ICN",
        "destination": "GUM",
        "route": "ICN-GUM",
        "airline": "LJ",
        "depFlight": "LJ915",
        "retFlight": "LJ918",
        "retDepartureTime": None,
        "defaultProductDays": 5,
        "fareRoute": "ICN-GUM-LJ",
        "enabled": True
    },
    {
        "key": "ICNGUM_LJ917_916",
        "rawKey": "ICNGUM_LJ917/6",
        "origin": "ICN",
        "destination": "GUM",
        "route": "ICN-GUM",
        "airline": "LJ",
        "depFlight": "LJ917",
        "retFlight": "LJ917",
        "retDepartureTime": None,
        "defaultProductDays": 5,
        "fareRoute": "ICN-GUM-LJ",
        "enabled": True
    },
    {
        "key": "ICNGUM_LJ917_918",
        "rawKey": "ICNGUM_LJ917/8",
        "origin": "ICN",
        "destination": "GUM",
        "route": "ICN-GUM",
        "airline": "LJ",
        "depFlight": "LJ917",
        "retFlight": "LJ918",
        "retDepartureTime": None,
        "defaultProductDays": 5,
        "fareRoute": "ICN-GUM-LJ",
        "enabled": True
    }
]


def ensure_hotels_manifest() -> Path:
    if HOTELS_MANIFEST_PATH.exists():
        return HOTELS_MANIFEST_PATH
    write_hotels_manifest(DEFAULT_INITIAL_MASTERS)
    return HOTELS_MANIFEST_PATH


def write_hotels_manifest(items: list[dict[str, object]], path: Path = HOTELS_MANIFEST_PATH) -> None:
    normalized = [normalize_flight_master_item(item) for item in items]
    payload = {
        "schema": "air-auto-lookup-flight-masters-v1",
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "flightMasters": normalized,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    atomic_write_text(path, json.dumps(payload, ensure_ascii=False, indent=2) + "\n")


def normalize_flight_master_item(item: dict[str, object]) -> dict[str, object]:
    origin = normalize_airport_code(item.get("origin"))
    destination = normalize_airport_code(item.get("destination"))
    airline = normalize_airline_code(item.get("airline"))
    if not origin or not destination or not airline:
        raise ValueError("출발지, 도착지, 항공사는 필수입니다.")

    dep_flight = normalize_manifest_flight(airline, item.get("depFlight"))
    ret_flight = normalize_manifest_flight(airline, item.get("retFlight"))
    if not dep_flight or not ret_flight:
        raise ValueError("출발편과 귀국편은 필수입니다.")

    route = str(item.get("route") or f"{origin}-{destination}").strip().upper()
    default_days = int(item.get("defaultProductDays") or 5)
    if default_days < 1:
        raise ValueError("상품일수는 1 이상이어야 합니다.")

    key = str(item.get("key") or "").strip().upper()
    raw_key = str(item.get("rawKey") or "").strip().upper()
    if not key:
        key = build_flight_master_key(origin, destination, airline, dep_flight, ret_flight)
    if not raw_key:
        raw_key = build_flight_master_raw_key(origin, destination, airline, dep_flight, ret_flight)

    ret_departure_time = item.get("retDepartureTime")
    ret_departure_time = None if ret_departure_time in (None, "") else str(ret_departure_time).strip()
    fare_route = str(item.get("fareRoute") or item.get("fare_route") or "").strip().upper()
    if not fare_route:
        fare_route = FARE_ROUTE_MAP.get(normalize_fare_route_key(raw_key)) or f"{route}-{airline}"

    return {
        "key": key,
        "rawKey": raw_key,
        "origin": origin,
        "destination": destination,
        "route": route,
        "airline": airline,
        "depFlight": dep_flight,
        "retFlight": ret_flight,
        "retDepartureTime": ret_departure_time,
        "defaultProductDays": default_days,
        "fareRoute": fare_route,
        "enabled": bool_from_manifest_value(item.get("enabled", True)),
    }


def normalize_airport_code(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())[:3]


def normalize_airline_code(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())[:2]


def normalize_manifest_flight(airline: str, value: object) -> str:
    text = re.sub(r"\s+", "", str(value or "").upper())
    text = re.sub(r"[^A-Z0-9]", "", text)
    while airline and text.startswith(f"{airline}{airline}"):
        text = text[len(airline) :]
    if airline and text.startswith(airline):
        return text
    if text and text[0].isdigit():
        text = f"{airline}{text}"
    return text


def bool_from_manifest_value(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"0", "false", "n", "no", "off", "disabled", "사용안함", "중지"}:
        return False
    return True


def flight_number_part(flight: str, airline: str) -> str:
    text = str(flight or "").upper()
    if text.startswith(airline):
        return text[len(airline) :]
    return text


def build_flight_master_key(origin: str, destination: str, airline: str, dep_flight: str, ret_flight: str) -> str:
    dep_number = flight_number_part(dep_flight, airline)
    ret_number = flight_number_part(ret_flight, airline)
    return re.sub(r"[^A-Z0-9_]", "_", f"{origin}{destination}_{airline}{dep_number}_{ret_number}".upper())


def build_flight_master_raw_key(origin: str, destination: str, airline: str, dep_flight: str, ret_flight: str) -> str:
    dep_number = flight_number_part(dep_flight, airline)
    ret_number = flight_number_part(ret_flight, airline)
    prefix_len = 0
    for left, right in zip(dep_number, ret_number):
        if left != right:
            break
        prefix_len += 1
    ret_suffix = ret_number[prefix_len:] or ret_number
    return f"{origin}{destination}_{airline}{dep_number}/{ret_suffix}".upper()


def flight_master_from_manifest_item(item: dict[str, object]) -> FlightMaster:
    normalized = normalize_flight_master_item(item)
    return FlightMaster(
        key=str(normalized["key"]),
        raw_key=str(normalized["rawKey"]),
        origin=str(normalized["origin"]),
        destination=str(normalized["destination"]),
        route=str(normalized["route"]),
        airline=str(normalized["airline"]),
        dep_flight=str(normalized["depFlight"]),
        ret_flight=str(normalized["retFlight"]),
        ret_departure_time=normalized.get("retDepartureTime"),
        default_product_days=int(normalized.get("defaultProductDays") or 5),
        fare_route=str(normalized.get("fareRoute") or ""),
        enabled=bool(normalized.get("enabled", True)),
    )


def _extract_js_array(text: str, var_name: str) -> str:
    marker = f"export const {var_name}"
    start = text.find(marker)
    if start < 0:
        raise ValueError(f"{var_name} export를 찾지 못했습니다.")
    bracket_start = text.find("[", start)
    if bracket_start < 0:
        raise ValueError(f"{var_name} 배열 시작을 찾지 못했습니다.")

    depth = 0
    quote = ""
    escape = False
    for idx in range(bracket_start, len(text)):
        ch = text[idx]
        if quote:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == quote:
                quote = ""
            continue
        if ch in ("'", '"'):
            quote = ch
            continue
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return text[bracket_start : idx + 1]
    raise ValueError(f"{var_name} 배열 끝을 찾지 못했습니다.")


def _js_object_array_to_json(array_text: str) -> str:
    text = re.sub(r"//.*", "", array_text)
    text = re.sub(r"([{,]\s*)([A-Za-z_][A-Za-z0-9_]*)\s*:", r'\1"\2":', text)
    text = re.sub(r",\s*([}\]])", r"\1", text)
    return text


def parse_iso_date(value: str) -> date:
    text = value.strip()
    if re.fullmatch(r"\d{8}", text):
        return datetime.strptime(text, "%Y%m%d").date()
    return datetime.strptime(text, "%Y-%m-%d").date()


def normalize_date_input(value: str) -> str:
    return parse_iso_date(value).isoformat()


def default_end_date(start: date | None = None) -> date:
    return (start or date.today()) + timedelta(days=359)


def iter_dates(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def date_count(start: date, end: date) -> int:
    return (end - start).days + 1


def topas_date(value: date) -> str:
    return f"{value.day:02d}{MONTH_CODES[value.month - 1]}"


def flight_number(flight: str, airline: str) -> str:
    compact_flight = str(flight).replace(" ", "").upper()
    compact_airline = str(airline).replace(" ", "").upper()
    if compact_flight.startswith(compact_airline):
        return compact_flight[len(compact_airline) :]
    return compact_flight


def topas_command(query_date: date, origin: str, destination: str, airline: str, flight: str) -> str:
    return f"AN{topas_date(query_date)}{origin}{destination}/A{airline}{flight_number(flight, airline)}"


def calculate_nights(product_days: int, ret_time: str | None, fallback_mode: str) -> tuple[int | None, str, str]:
    if product_days < 2:
        return None, "상품일수오류", "상품일수는 2일 이상이어야 합니다."

    parsed = _parse_time(ret_time)
    if parsed is not None:
        minutes = parsed
        if 0 <= minutes < 360:
            return product_days - 1, "00시 이후", f"귀국편 시간 {ret_time}: 상품일수 - 1"
        if 360 <= minutes <= 1439:
            return product_days - 2, "24시 이전", f"귀국편 시간 {ret_time}: 상품일수 - 2"
        return None, "시간확인필요", f"귀국편 시간 {ret_time}: MVP 자동분류 범위 밖"

    if fallback_mode == "before_midnight":
        return product_days - 2, "24시 이전 기본값", "귀국편 시간 미입력: 상품일수 - 2 기본 적용"
    if fallback_mode == "after_midnight":
        return product_days - 1, "00시 이후 기본값", "귀국편 시간 미입력: 상품일수 - 1 기본 적용"
    return None, "시간확인필요", "귀국편 시간 미입력: 수동확인 필요"


def return_candidate_dates(dep_date: date, product_days: int) -> list[tuple[int, date]]:
    if product_days < 2:
        return []
    candidates = []
    for nights in (product_days - 2, product_days - 1):
        if nights >= 0:
            candidates.append((nights, dep_date + timedelta(days=nights)))
    return candidates


def _parse_time(value: str | None) -> int | None:
    if not value:
        return None
    match = re.fullmatch(r"\s*(\d{1,2}):?(\d{2})\s*", str(value))
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return hour * 60 + minute


def _parse_day_offset(value: object) -> int:
    text = str(value or "").strip()
    if not text:
        return 0
    try:
        return int(text)
    except ValueError:
        return 0


def _return_candidate_arrival_date(item: dict[str, object]) -> date | None:
    candidate_date_text = str(item.get("candidateReturnDate") or "").strip()
    if not candidate_date_text:
        return None
    try:
        candidate_date = parse_iso_date(candidate_date_text)
    except ValueError:
        return None
    offset: object = item.get("arrivalDayOffset")
    if offset in (None, ""):
        matches = item.get("matches", [])
        if isinstance(matches, list) and matches and isinstance(matches[0], dict):
            offset = matches[0].get("arriveDayOffset")
    return candidate_date + timedelta(days=_parse_day_offset(offset))


def _return_candidate_arrival_time(item: dict[str, object]) -> str:
    detected_arrival_time = str(item.get("detectedArrivalTime") or "").strip()
    if detected_arrival_time:
        return detected_arrival_time
    matches = item.get("matches", [])
    if isinstance(matches, list) and matches and isinstance(matches[0], dict):
        return str(matches[0].get("arriveTime") or "").strip()
    return ""


def build_plan_rows(
    masters: list[FlightMaster],
    start: date,
    end: date,
    product_days: int,
    fallback_mode: str,
    detected_return_times: dict[tuple[str, str], str] | None = None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    detected_return_times = detected_return_times or {}
    for master in masters:
        for dep_date in iter_dates(start, end):
            row_product_days = product_days or master.default_product_days
            detected_ret_time = detected_return_times.get((master.key, dep_date.isoformat()))
            ret_time_basis = detected_ret_time or master.ret_departure_time
            if ret_time_basis:
                nights, night_type, night_basis = calculate_nights(
                    row_product_days,
                    ret_time_basis,
                    fallback_mode,
                )
            else:
                nights = None
                night_type = "귀국시간감지필요"
                night_basis = "귀국편 후보일을 조회한 뒤 TOPAS 실제 출발시간으로 박수 확정"

            ret_date = dep_date + timedelta(days=nights) if nights is not None else None
            dep_command = topas_command(
                dep_date,
                master.origin,
                master.destination,
                master.airline,
                master.dep_flight,
            )
            ret_command = (
                topas_command(
                    ret_date,
                    master.destination,
                    master.origin,
                    master.airline,
                    master.ret_flight,
                )
                if ret_date is not None
                else ""
            )
            candidate_commands = [
                f"{candidate_nights}박:{topas_command(candidate_date, master.destination, master.origin, master.airline, master.ret_flight)}"
                for candidate_nights, candidate_date in return_candidate_dates(dep_date, row_product_days)
            ]
            rows.append(
                {
                    "route_key": master.key,
                    "raw_key": master.raw_key,
                    "route": master.route,
                    "origin": master.origin,
                    "destination": master.destination,
                    "airline": master.airline,
                    "dep_flight": master.dep_flight,
                    "ret_flight": master.ret_flight,
                    "base_departure_date": dep_date.isoformat(),
                    "product_days": row_product_days,
                    "ret_departure_time": ret_time_basis or "",
                    "nights": nights if nights is not None else "",
                    "night_type": night_type,
                    "night_basis": night_basis,
                    "dep_query_date": dep_date.isoformat(),
                    "dep_command": dep_command,
                    "ret_query_date": ret_date.isoformat() if ret_date else "",
                    "ret_command": ret_command,
                    "ret_candidate_commands": " | ".join(candidate_commands) if ret_date is None else "",
                    "dep_status": "조회전",
                    "ret_status": "조회전" if ret_date else "귀국시간감지필요",
                    "dep_classes": "",
                    "ret_classes": "",
                    "operating": "",
                    "round_trip_fare": "",
                    "fare_status": "요금계산전",
                    "raw_departure_file": "",
                    "raw_return_file": "",
                }
            )
    return rows


def build_command_plan(
    run_id: str,
    masters: list[FlightMaster],
    start: date,
    end: date,
    product_days: int,
) -> dict[str, object]:
    departure_commands: list[dict[str, object]] = []
    return_candidate_commands: list[dict[str, object]] = []

    for master in masters:
        fare_route = fare_route_for_master(master)
        for dep_date in iter_dates(start, end):
            row_product_days = product_days or master.default_product_days
            dep_command = topas_command(
                dep_date,
                master.origin,
                master.destination,
                master.airline,
                master.dep_flight,
            )
            departure_commands.append(
                {
                    "id": f"{master.key}:{dep_date.isoformat()}:departure",
                    "routeKey": master.key,
                    "rawKey": master.raw_key,
                    "route": master.route,
                    "fareRoute": fare_route,
                    "baseDepartureDate": dep_date.isoformat(),
                    "queryDate": dep_date.isoformat(),
                    "direction": "departure",
                    "origin": master.origin,
                    "destination": master.destination,
                    "airline": master.airline,
                    "flight": master.dep_flight,
                    "command": dep_command,
                    "rawFile": str(raw_file_path(run_id, master.key, dep_date, "departure")),
                    "status": "planned",
                }
            )

            for candidate_nights, candidate_date in return_candidate_dates(dep_date, row_product_days):
                ret_command = topas_command(
                    candidate_date,
                    master.destination,
                    master.origin,
                    master.airline,
                    master.ret_flight,
                )
                return_candidate_commands.append(
                    {
                        "id": f"{master.key}:{dep_date.isoformat()}:return:{candidate_nights}n",
                        "routeKey": master.key,
                        "rawKey": master.raw_key,
                        "route": master.route,
                        "fareRoute": fare_route,
                        "baseDepartureDate": dep_date.isoformat(),
                        "productDays": row_product_days,
                        "candidateNights": candidate_nights,
                        "candidateReturnDate": candidate_date.isoformat(),
                        "queryDate": candidate_date.isoformat(),
                        "direction": "return",
                        "origin": master.destination,
                        "destination": master.origin,
                        "airline": master.airline,
                        "flight": master.ret_flight,
                        "command": ret_command,
                        "rawFile": str(return_raw_file_path(run_id, master.key, candidate_date)),
                        "status": "planned",
                    }
                )

    raw_root = RAW_DIR / run_id
    raw_store = raw_root / RAW_STORE_FILENAME
    return {
        "runId": run_id,
        "generatedAt": now_iso(),
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "dateCount": date_count(start, end),
        "productDays": product_days,
        "departureCommands": departure_commands,
        "returnCandidateCommands": return_candidate_commands,
        "rawPathPolicy": {
            "storage": "sqlite",
            "root": str(raw_root),
            "rawStore": str(raw_store),
            "departurePattern": str(RAW_DIR / run_id / "{routeKey}" / "departure" / "{YYYY-MM-DD}_departure.txt"),
            "returnPattern": str(RAW_DIR / run_id / "{routeKey}" / "return" / "{YYYY-MM-DD}_return.txt"),
            "txtPathsRetainedForExportCompatibility": True,
            "dryRunCreatesEmptyRawFiles": False,
        },
    }


def raw_file_path(
    run_id: str,
    route_key: str,
    base_departure_date: date,
    direction: str,
    candidate_nights: int | None = None,
) -> Path:
    if direction == "departure":
        filename = f"{base_departure_date.isoformat()}_departure.txt"
    elif direction == "return" and candidate_nights is not None:
        filename = f"{base_departure_date.isoformat()}_return_{candidate_nights}n.txt"
    else:
        raise ValueError(f"지원하지 않는 원문 경로 유형입니다: {direction}")
    return RAW_DIR / run_id / route_key / direction / filename


def return_raw_file_path(run_id: str, route_key: str, candidate_return_date: date) -> Path:
    return RAW_DIR / run_id / route_key / "return" / f"{candidate_return_date.isoformat()}_return.txt"


def ensure_raw_plan_dirs(command_plan: dict[str, object]) -> None:
    raw_policy = command_plan.get("rawPathPolicy", {}) if isinstance(command_plan.get("rawPathPolicy"), dict) else {}
    raw_store = str(raw_policy.get("rawStore") or "").strip()
    if raw_store:
        Path(raw_store).parent.mkdir(parents=True, exist_ok=True)
        return
    root = str(raw_policy.get("root") or "").strip()
    if root:
        Path(root).mkdir(parents=True, exist_ok=True)


def write_command_plan(run_doc: dict[str, object], command_plan: dict[str, object]) -> Path:
    log_dir = Path(str(run_doc["logDir"]))
    path = log_dir / "command-plan.json"
    payload = json.dumps(command_plan, ensure_ascii=False, indent=2)
    atomic_write_text(path, payload + "\n")
    ensure_raw_plan_dirs(command_plan)
    departure_count = len(command_plan.get("departureCommands", []))
    return_count = len(command_plan.get("returnCandidateCommands", []))
    run_id = str(run_doc["runId"])
    raw_policy = command_plan.get("rawPathPolicy", {}) if isinstance(command_plan.get("rawPathPolicy"), dict) else {}
    raw_root = str(raw_policy.get("root") or RAW_DIR / run_id)
    raw_store = str(raw_policy.get("rawStore") or Path(raw_root) / RAW_STORE_FILENAME)
    run_doc["commandPlan"] = str(path)
    run_doc["rawRoot"] = raw_root
    run_doc["rawStore"] = raw_store
    run_doc["departureCommandCount"] = departure_count
    run_doc["returnCandidateCommandCount"] = return_count
    write_run_json(run_doc)
    append_run_event(
        run_doc,
        "command_plan_created",
        commandPlan=str(path),
        departureCommandCount=departure_count,
        returnCandidateCommandCount=return_count,
    )
    return path


def _command_plan_items(command_plan: dict[str, object]):
    for key in ("departureCommands", "returnCandidateCommands"):
        items = command_plan.get(key, [])
        if isinstance(items, list):
            yield from (item for item in items if isinstance(item, dict))


def command_raw_files(command: dict[str, object]) -> list[Path]:
    raw_files: list[Path] = []
    raw_file_values = command.get("rawFiles")
    if isinstance(raw_file_values, list):
        raw_files.extend(Path(str(value)) for value in raw_file_values if str(value or "").strip())
    raw_file = str(command.get("rawFile") or "").strip()
    if raw_file:
        raw_files.append(Path(raw_file))
    selected: list[Path] = []
    seen: set[str] = set()
    for raw_path in raw_files:
        key = str(raw_path)
        if key in seen:
            continue
        seen.add(key)
        selected.append(raw_path)
    return selected


def command_identity(command: dict[str, object]) -> str:
    parts = [
        str(command.get("direction") or ""),
        str(command.get("command") or ""),
        str(command.get("queryDate") or ""),
        str(command.get("origin") or ""),
        str(command.get("destination") or ""),
        str(command.get("airline") or ""),
        str(command.get("flight") or ""),
    ]
    identity = "|".join(part.strip().upper() for part in parts if part.strip())
    return identity or str(command.get("rawFile") or command.get("id") or "")


def normalize_topas_flight_number(value: object, airline: object = "") -> str:
    text = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    airline_text = re.sub(r"[^A-Z0-9]", "", str(airline or "").upper())
    if airline_text and text.startswith(airline_text):
        text = text[len(airline_text) :]
    digits = re.sub(r"\D", "", text)
    return str(int(digits)) if digits else ""


def topas_command_signature_from_match(match: re.Match[str]) -> tuple[str, str, str, str, str]:
    airline = match.group("airline").upper()
    return (
        match.group("date").upper().zfill(5),
        match.group("origin").upper(),
        match.group("destination").upper(),
        airline,
        normalize_topas_flight_number(match.group("flight"), airline),
    )


def topas_command_signature_from_text(value: object) -> tuple[str, str, str, str, str] | None:
    match = TOPAS_FULL_COMMAND_RE.search(str(value or "").upper())
    return topas_command_signature_from_match(match) if match else None


def topas_command_signature_from_row(command: dict[str, object]) -> tuple[str, str, str, str, str] | None:
    command_signature = topas_command_signature_from_text(command.get("command"))
    if command_signature:
        return command_signature
    query_date = str(command.get("queryDate") or "").strip()
    airline = str(command.get("airline") or "").strip().upper()
    if not query_date or not airline:
        return None
    try:
        parsed_date = parse_iso_date(query_date)
    except Exception:
        return None
    return (
        f"{parsed_date.day:02d}{MONTH_NAMES[parsed_date.month - 1]}",
        str(command.get("origin") or "").strip().upper(),
        str(command.get("destination") or "").strip().upper(),
        airline,
        normalize_topas_flight_number(command.get("flight"), airline),
    )


def first_response_command_signature(text: str) -> tuple[str, str, str, str, str] | None:
    for line in str(text or "").replace("\r\n", "\n").splitlines():
        if not line.strip():
            continue
        return topas_command_signature_from_text(line)
    return topas_command_signature_from_text(text)


def command_response_mismatch(command: dict[str, object], raw_text: str) -> bool:
    expected = topas_command_signature_from_row(command)
    actual = first_response_command_signature(raw_text)
    return bool(expected and actual and expected != actual)


def command_mismatch_reason(command: dict[str, object], raw_text: str) -> str:
    expected = str(command.get("command") or "").strip()
    first_line = ""
    for line in str(raw_text or "").replace("\r\n", "\n").splitlines():
        if line.strip():
            first_line = line.strip()
            break
    if first_line:
        return f"수집 원문 명령 불일치: 요청 {expected} / 원문 {first_line}"
    return f"수집 원문 명령 불일치: 요청 {expected}"


def raw_text_state(text: str) -> str:
    stripped = str(text or "").strip()
    if not stripped:
        return "empty_or_truncated"
    upper = stripped.upper()
    first_line = upper.splitlines()[0] if upper.splitlines() else ""
    if first_line.startswith("TOPAS_COLLECT_ERROR") or first_line.startswith("ERROR"):
        return "collect_error"
    if "NO FLIGHT" in upper:
        return "no_flight"
    if any(marker in upper for marker in RAW_COMPLETE_MARKERS):
        return "normal_raw"
    return "empty_or_truncated"


def raw_store_path(log_dir: str | Path) -> Path:
    log_path = Path(log_dir)
    run_path = log_path / "run.json"
    if run_path.exists():
        try:
            run_doc = json.loads(run_path.read_text(encoding="utf-8"))
        except Exception:
            run_doc = {}
        if isinstance(run_doc, dict):
            raw_store = str(run_doc.get("rawStore") or "").strip()
            if raw_store:
                return Path(raw_store)
            raw_root = str(run_doc.get("rawRoot") or "").strip()
            if raw_root:
                return Path(raw_root) / RAW_STORE_FILENAME
    return log_path / RAW_STORE_FILENAME


def raw_record_for_command(log_dir: str | Path, command: dict[str, object]) -> dict[str, object] | None:
    path = raw_store_path(log_dir)
    if not path.exists():
        return None
    try:
        conn = sqlite3.connect(path)
        conn.row_factory = sqlite3.Row
        try:
            found = conn.execute(
                "SELECT raw_text, status, error FROM raw_responses WHERE command_identity = ?",
                (command_identity(command),),
            ).fetchone()
        finally:
            conn.close()
    except Exception:
        return None
    return dict(found) if found is not None else None


def load_raw_store_records(log_dir: str | Path) -> dict[str, dict[str, object]]:
    path = raw_store_path(log_dir)
    if not path.exists():
        return {}
    try:
        conn = sqlite3.connect(path)
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute("SELECT command_identity, raw_text, status, error FROM raw_responses").fetchall()
        finally:
            conn.close()
    except Exception:
        return {}
    return {str(row["command_identity"]): dict(row) for row in rows}


def raw_record_from_cache(
    raw_records: dict[str, dict[str, object]] | None,
    command: dict[str, object],
) -> dict[str, object] | None:
    if not raw_records:
        return None
    return raw_records.get(command_identity(command))


def raw_response_state_from_record(
    record: dict[str, object] | None,
    raw_path: Path,
    command: dict[str, object] | None = None,
) -> str:
    if record is not None:
        if command is not None and command_response_mismatch(command, str(record.get("raw_text") or "")):
            return "command_mismatch"
        status = str(record.get("status") or "").strip()
        return status or raw_text_state(str(record.get("raw_text") or ""))
    return raw_file_state(raw_path)


def read_raw_response_text_from_record(
    record: dict[str, object] | None,
    raw_path: Path,
) -> tuple[str, bool, str, str]:
    if record is not None:
        return str(record.get("raw_text") or "").rstrip(), True, "", "sqlite"
    if not raw_path.exists():
        return "", False, "", ""
    try:
        return raw_path.read_text(encoding="utf-8", errors="replace").rstrip(), True, "", "txt"
    except Exception as exc:
        return f"[RAW READ ERROR] {exc}", True, str(exc), "txt"


def load_retry_state(log_dir: Path) -> dict[str, object]:
    path = log_dir / RETRY_STATE_FILE
    if not path.exists():
        return {"version": 1, "commands": {}}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"version": 1, "commands": {}}
    if not isinstance(payload, dict):
        return {"version": 1, "commands": {}}
    if not isinstance(payload.get("commands"), dict):
        payload["commands"] = {}
    payload.setdefault("version", 1)
    return payload


def retry_count_for_command(retry_state: dict[str, object], command: dict[str, object]) -> int:
    commands = retry_state.get("commands", {})
    if not isinstance(commands, dict):
        return 0
    record = commands.get(command_identity(command))
    if not isinstance(record, dict):
        return 0
    try:
        return int(record.get("retryCount") or 0)
    except (TypeError, ValueError):
        return 0


def retry_exhausted_error_for_raw(retry_state: dict[str, object], raw_file: Path) -> str:
    commands = retry_state.get("commands", {})
    if not isinstance(commands, dict):
        return ""
    raw_text = str(raw_file)
    for record in commands.values():
        if not isinstance(record, dict) or not record.get("retryExhausted"):
            continue
        raw_files = record.get("rawFiles", [])
        if isinstance(raw_files, list) and raw_text in {str(value) for value in raw_files}:
            return str(record.get("lastError") or "TOPAS 수집 오류가 재시도 한도를 초과했습니다.")
    return ""


def raw_file_state(path: Path) -> str:
    if not path.exists():
        return "missing"
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return "collect_error"
    stripped = text.strip()
    if not stripped:
        return "empty_or_truncated"
    upper = stripped.upper()
    first_line = upper.splitlines()[0] if upper.splitlines() else ""
    if first_line.startswith("TOPAS_COLLECT_ERROR") or first_line.startswith("ERROR"):
        return "collect_error"
    if "NO FLIGHT" in upper:
        return "no_flight"
    if any(marker in upper for marker in RAW_COMPLETE_MARKERS):
        return "normal_raw"
    return "empty_or_truncated"


def raw_response_state(log_dir: str | Path, command: dict[str, object], raw_path: Path) -> str:
    record = raw_record_for_command(log_dir, command)
    return raw_response_state_from_record(record, raw_path, command)


def read_raw_response_text(log_dir: str | Path, command: dict[str, object], raw_path: Path) -> tuple[str, bool, str]:
    record = raw_record_for_command(log_dir, command)
    raw_text, exists, read_error, _source = read_raw_response_text_from_record(record, raw_path)
    return raw_text, exists, read_error


def collect_pending_summary(log_dir: str | Path, max_retries: int = DEFAULT_MAX_COLLECT_ERROR_RETRIES) -> dict[str, object]:
    log_path = Path(log_dir)
    command_plan_path = log_path / "command-plan.json"
    if not command_plan_path.exists():
        return {"actionablePendingCommands": 0, "missing": 0, "retryableErrors": 0}
    command_plan = json.loads(command_plan_path.read_text(encoding="utf-8"))
    rows = [row for row in _command_plan_items(command_plan)]
    retry_state = load_retry_state(log_path)
    raw_records = load_raw_store_records(log_path)
    counts = {
        "missing": 0,
        "normal_raw": 0,
        "no_flight": 0,
        "collect_error_retryable": 0,
        "empty_or_truncated_retryable": 0,
        "command_mismatch_retryable": 0,
        "retry_exhausted": 0,
    }
    actionable_identities: set[str] = set()
    unique_raw_files: set[str] = set()
    unique_commands: set[str] = set()
    for command in rows:
        identity = command_identity(command)
        unique_commands.add(identity)
        retries = retry_count_for_command(retry_state, command)
        for raw_path in command_raw_files(command):
            unique_raw_files.add(str(raw_path))
            state = raw_response_state_from_record(raw_record_from_cache(raw_records, command), raw_path, command)
            if state == "missing":
                counts["missing"] += 1
                actionable_identities.add(identity)
            elif state in {"collect_error", "empty_or_truncated", "command_mismatch"}:
                if retries < max_retries:
                    counts[f"{state}_retryable"] += 1
                    actionable_identities.add(identity)
                else:
                    counts["retry_exhausted"] += 1
            else:
                counts[state] += 1
    return {
        "logicalRows": len(rows),
        "uniqueTopasCommands": len(unique_commands),
        "uniqueRawFiles": len(unique_raw_files),
        "actionablePendingCommands": len(actionable_identities),
        **counts,
    }


def has_actionable_pending(log_dir: str | Path, max_retries: int = DEFAULT_MAX_COLLECT_ERROR_RETRIES) -> bool:
    return int(collect_pending_summary(log_dir, max_retries).get("actionablePendingCommands") or 0) > 0


def generate_combined_raw_views(log_dir: str | Path) -> dict[str, Path]:
    log_path = Path(log_dir)
    run_json_path = log_path / "run.json"
    command_plan_path = log_path / "command-plan.json"
    if not run_json_path.exists() or not command_plan_path.exists():
        return {}

    run_doc = json.loads(run_json_path.read_text(encoding="utf-8"))
    command_plan = json.loads(command_plan_path.read_text(encoding="utf-8"))
    run_id = str(run_doc.get("runId") or command_plan.get("runId") or log_path.name)
    raw_policy = command_plan.get("rawPathPolicy", {}) if isinstance(command_plan.get("rawPathPolicy"), dict) else {}
    raw_root = Path(str(run_doc.get("rawRoot") or raw_policy.get("root") or RAW_DIR / run_id))
    combined_dir = raw_root / "_combined"
    combined_dir.mkdir(parents=True, exist_ok=True)

    grouped: dict[str, dict[str, object]] = {}
    for command in _command_plan_items(command_plan):
        for raw_path in command_raw_files(command):
            key = str(raw_path)
            group = grouped.setdefault(key, {"rawFile": raw_path, "commands": []})
            commands = group["commands"]
            if isinstance(commands, list):
                commands.append(command)

    def command_values(commands: list[dict[str, object]], field: str) -> list[str]:
        return sorted({str(command.get(field) or "") for command in commands if str(command.get(field) or "").strip()})

    def sort_key(group: dict[str, object]) -> tuple[object, ...]:
        commands = group.get("commands", [])
        first = commands[0] if isinstance(commands, list) and commands else {}
        direction = str(first.get("direction") or "")
        direction_order = 0 if direction == "departure" else 1 if direction == "return" else 2
        return (
            direction_order,
            str(first.get("routeKey") or ""),
            str(first.get("queryDate") or ""),
            str(first.get("command") or ""),
            str(group.get("rawFile") or ""),
        )

    generated_at = now_iso()
    all_path = combined_dir / "all_raw.txt"
    departure_path = combined_dir / "departure_raw.txt"
    return_path = combined_dir / "return_raw.txt"
    raw_files_log_path = log_path / "raw-files.jsonl"
    index_path = log_path / "raw-view-index.json"
    raw_records = load_raw_store_records(log_path)

    header = [
        "# 항공자동조회 통합 RAW VIEW",
        f"generatedAt: {generated_at}",
        f"runId: {run_id}",
        f"rawRoot: {raw_root}",
        f"recordCount: {len(grouped)}",
        "",
    ]
    all_lines: list[str] = list(header)
    direction_lines: dict[str, list[str]] = {
        "departure": [*header, "direction: departure", ""],
        "return": [*header, "direction: return", ""],
    }
    records: list[dict[str, object]] = []
    state_counts: dict[str, int] = {}

    for index, group in enumerate(sorted(grouped.values(), key=sort_key), start=1):
        raw_path = group["rawFile"]
        commands = group.get("commands", [])
        command_list = commands if isinstance(commands, list) else []
        first = command_list[0] if command_list else {}
        assert isinstance(raw_path, Path)
        raw_record = raw_record_from_cache(raw_records, first)
        state = raw_response_state_from_record(raw_record, raw_path, first)
        state_counts[state] = state_counts.get(state, 0) + 1
        raw_text, exists, read_error, raw_source = read_raw_response_text_from_record(raw_record, raw_path)
        byte_length = len(raw_text.encode("utf-8")) if exists else 0
        if not exists:
            raw_text = "[MISSING RAW FILE]"
            raw_source = ""

        base_departure_dates = command_values(command_list, "baseDepartureDate")
        command_ids = command_values(command_list, "id")
        candidate_nights = command_values(command_list, "candidateNights")
        candidate_return_dates = command_values(command_list, "candidateReturnDate")
        metadata = {
            "index": index,
            "runId": run_id,
            "direction": str(first.get("direction") or ""),
            "routeKey": str(first.get("routeKey") or ""),
            "rawKey": str(first.get("rawKey") or ""),
            "route": str(first.get("route") or ""),
            "fareRoute": str(first.get("fareRoute") or ""),
            "baseDepartureDates": base_departure_dates,
            "logicalRowCount": len(command_list),
            "commandIds": command_ids,
            "candidateNights": candidate_nights,
            "candidateReturnDates": candidate_return_dates,
            "queryDate": str(first.get("queryDate") or ""),
            "origin": str(first.get("origin") or ""),
            "destination": str(first.get("destination") or ""),
            "airline": str(first.get("airline") or ""),
            "flight": str(first.get("flight") or ""),
            "command": str(first.get("command") or ""),
            "commandIdentity": command_identity(first) if first else "",
            "rawFile": str(raw_path),
            "rawStore": str(raw_store_path(log_path)),
            "rawSource": raw_source,
            "exists": exists,
            "state": state,
            "byteLength": byte_length,
            "readError": read_error,
        }
        records.append(metadata)

        section = [
            "=" * 96,
            f"[{index:04d}] {metadata['direction']} {metadata['routeKey']} {metadata['queryDate']} {metadata['command']}",
            f"state: {state}",
            f"rawFile: {raw_path}",
            f"rawSource: {raw_source or 'missing'}",
            f"logicalRowCount: {len(command_list)}",
            f"baseDepartureDates: {', '.join(base_departure_dates) if base_departure_dates else ''}",
        ]
        if candidate_nights:
            section.append(f"candidateNights: {', '.join(candidate_nights)}")
        if candidate_return_dates:
            section.append(f"candidateReturnDates: {', '.join(candidate_return_dates)}")
        section.extend(["-" * 96, raw_text, ""])
        all_lines.extend(section)
        direction = str(metadata["direction"])
        if direction in direction_lines:
            direction_lines[direction].extend(section)

    all_path.write_text("\n".join(all_lines).rstrip() + "\n", encoding="utf-8")
    departure_path.write_text("\n".join(direction_lines["departure"]).rstrip() + "\n", encoding="utf-8")
    return_path.write_text("\n".join(direction_lines["return"]).rstrip() + "\n", encoding="utf-8")
    raw_files_log_path.write_text(
        "".join(json.dumps(record, ensure_ascii=False) + "\n" for record in records),
        encoding="utf-8",
    )
    index_payload = {
        "generatedAt": generated_at,
        "runId": run_id,
        "rawRoot": str(raw_root.resolve()),
        "recordCount": len(records),
        "stateCounts": state_counts,
        "allRaw": str(all_path.resolve()),
        "departureRaw": str(departure_path.resolve()),
        "returnRaw": str(return_path.resolve()),
        "rawFilesLog": str(raw_files_log_path.resolve()),
    }
    index_path.write_text(json.dumps(index_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    run_doc["rawView"] = {
        **index_payload,
        "index": str(index_path),
    }
    atomic_write_text(run_json_path, json.dumps(run_doc, ensure_ascii=False, indent=2) + "\n")
    append_run_event(
        run_doc,
        "combined_raw_view_generated",
        recordCount=len(records),
        stateCounts=state_counts,
        allRaw=str(all_path.resolve()),
        rawFilesLog=str(raw_files_log_path.resolve()),
    )
    return {
        "allRaw": all_path.resolve(),
        "departureRaw": departure_path.resolve(),
        "returnRaw": return_path.resolve(),
        "rawFilesLog": raw_files_log_path.resolve(),
        "rawViewIndex": index_path.resolve(),
    }


def parse_collect_error_raw(text: str) -> dict[str, object] | None:
    lines = str(text or "").replace("\r\n", "\n").splitlines()
    if not lines:
        return None
    first_line = lines[0].upper()
    if not (first_line.startswith("TOPAS_COLLECT_ERROR") or first_line.startswith("ERROR")):
        return None
    error = ""
    command = ""
    for line in lines[1:]:
        if line.upper().startswith("COMMAND:"):
            command = line.split(":", 1)[1].strip()
        elif line.upper().startswith("ERROR:"):
            error = line.split(":", 1)[1].strip()
    return {
        "status": "collect_failed",
        "command": command,
        "error": error or "TOPAS 수집 중 오류가 발생했습니다.",
        "matches": [],
    }


def parse_return_time_from_topas(text: str, airline: str, ret_flight: str) -> dict[str, object]:
    parsed = parse_topas_flight_result(text, airline, ret_flight)
    if parsed.get("status") == "collect_failed":
        return {
            "flight": ret_flight,
            "detectedReturnTime": "",
            "status": "collect_failed",
            "error": parsed.get("error", ""),
            "matches": [],
        }
    matches = parsed.get("matches", [])
    unique_schedules = sorted(
        {
            (
                str(item.get("departTime") or ""),
                str(item.get("arriveTime") or ""),
                _parse_day_offset(item.get("arriveDayOffset")),
            )
            for item in matches
            if item.get("departTime")
        }
    )
    if len(unique_schedules) == 1:
        depart_time, arrive_time, arrive_day_offset = unique_schedules[0]
        return {
            "flight": ret_flight,
            "detectedReturnTime": depart_time,
            "detectedArrivalTime": arrive_time,
            "arrivalDayOffset": arrive_day_offset,
            "status": "detected",
            "matches": matches,
        }
    if len(unique_schedules) > 1:
        return {
            "flight": ret_flight,
            "detectedReturnTime": "",
            "detectedArrivalTime": "",
            "arrivalDayOffset": "",
            "status": "ambiguous",
            "matches": matches,
        }

    upper_text = str(text or "").upper()
    if "NO FLIGHT" in upper_text:
        status = "no_flight"
    else:
        status = "parse_failed"
    return {
        "flight": ret_flight,
        "detectedReturnTime": "",
        "detectedArrivalTime": "",
        "arrivalDayOffset": "",
        "status": status,
        "matches": [],
    }


def parse_topas_flight_result(text: str, airline: str, flight: str) -> dict[str, object]:
    airline = str(airline).strip().upper()
    flight = str(flight).strip().upper()
    collect_error = parse_collect_error_raw(text)
    if collect_error:
        return {"flight": flight, **collect_error}
    target_number = flight_number(flight, airline).lstrip("0") or "0"
    matches: list[dict[str, object]] = []

    lines = str(text or "").replace("\r\n", "\n").splitlines()
    for index, line in enumerate(lines):
        match = TOPAS_FLIGHT_LINE_RE.match(line)
        if not match:
            continue
        line_airline = match.group("airline").upper()
        line_number = match.group("flight").lstrip("0") or "0"
        if line_airline != airline or line_number != target_number:
            continue
        class_text = match.group("class_text") or ""
        if index + 1 < len(lines) and "/" not in lines[index + 1]:
            class_text = f"{class_text} {lines[index + 1]}"
        classes = parse_class_tokens(class_text)
        arrive_day_offset_text = match.group("arrive_day_offset") or ""
        arrive_day_offset = _parse_day_offset(arrive_day_offset_text)
        matches.append(
            {
                "flight": f"{line_airline}{line_number}",
                "departTime": format_hhmm(match.group("depart_time")),
                "arriveTime": format_hhmm(match.group("arrive_time")),
                "arriveDayOffset": arrive_day_offset,
                "arriveDayOffsetText": arrive_day_offset_text,
                "classes": classes,
                "rawLine": line.strip(),
            }
        )

    if matches:
        return {
            "flight": flight,
            "status": "detected" if len(matches) == 1 else "ambiguous",
            "matches": matches,
        }

    upper_text = str(text or "").upper()
    status = "no_flight" if "NO FLIGHT" in upper_text else "parse_failed"
    return {
        "flight": flight,
        "status": status,
        "matches": [],
    }


def parse_class_tokens(text: str) -> dict[str, str]:
    return {code.upper(): value.upper() for code, value in TOPAS_CLASS_RE.findall(str(text or ""))}


def format_hhmm(value: str) -> str:
    text = str(value).strip().zfill(4)
    return f"{text[:2]}:{text[2:]}"


def confirm_return_nights(
    base_departure_date: str,
    route_key: str,
    product_days: int,
    candidate_results: list[dict[str, object]],
) -> dict[str, object]:
    confirmed: list[dict[str, object]] = []
    statuses = {str(item.get("status", "")) for item in candidate_results}
    base_date = parse_iso_date(base_departure_date)
    target_arrival_date = base_date + timedelta(days=product_days - 1)

    for item in candidate_results:
        if item.get("status") != "detected":
            continue
        arrival_date = _return_candidate_arrival_date(item)
        if arrival_date is None or arrival_date != target_arrival_date:
            continue
        candidate_nights = int(item.get("candidateNights", -1))
        arrival_time = _return_candidate_arrival_time(item)
        confirmed.append(
            {
                **item,
                "detectedArrivalDate": arrival_date.isoformat(),
                "detectedArrivalTime": arrival_time,
                "nights": candidate_nights if candidate_nights >= 0 else "",
                "scheduleDays": product_days,
                "targetArrivalDate": target_arrival_date.isoformat(),
                "nightBasis": (
                    f"한국 도착일 {arrival_date.isoformat()}: "
                    f"{base_departure_date} 출발 {product_days}일 일정과 일치"
                ),
            }
        )

    if len(confirmed) == 1:
        selected = confirmed[0]
        return {
            "baseDepartureDate": base_departure_date,
            "routeKey": route_key,
            "detectedReturnDate": selected.get("candidateReturnDate", ""),
            "detectedReturnTime": selected.get("detectedReturnTime", ""),
            "detectedArrivalDate": selected.get("detectedArrivalDate", ""),
            "detectedArrivalTime": selected.get("detectedArrivalTime", ""),
            "targetArrivalDate": target_arrival_date.isoformat(),
            "scheduleDays": product_days,
            "nights": selected.get("nights", ""),
            "status": "confirmed",
            "confirmationMode": "arrival_date_matched",
            "basis": selected.get("nightBasis", ""),
            "candidates": candidate_results,
        }
    if len(confirmed) > 1:
        status = "ambiguous"
    elif "raw_missing" in statuses:
        status = "raw_missing"
    elif "collect_failed" in statuses:
        status = "collect_failed"
    elif "command_mismatch" in statuses:
        status = "command_mismatch"
    elif statuses and statuses <= {"no_flight"}:
        status = "no_flight"
    elif "ambiguous" in statuses:
        status = "ambiguous"
    elif any(item.get("status") == "detected" for item in candidate_results):
        status = "schedule_mismatch"
    else:
        status = "parse_failed"

    mismatch_reasons = [
        str(item.get("error") or "").strip()
        for item in candidate_results
        if item.get("status") == "command_mismatch" and str(item.get("error") or "").strip()
    ]
    if status == "command_mismatch" and mismatch_reasons:
        basis = mismatch_reasons[0]
    elif status == "schedule_mismatch":
        detected_summaries: list[str] = []
        for item in candidate_results:
            if item.get("status") != "detected":
                continue
            arrival_date = _return_candidate_arrival_date(item)
            arrival_time = _return_candidate_arrival_time(item)
            if arrival_date is None:
                continue
            detected_summaries.append(
                f"{item.get('candidateReturnDate', '')} 출발 -> "
                f"{arrival_date.isoformat()} {arrival_time} 한국도착"
            )
        basis = (
            f"목표 한국 도착일 {target_arrival_date.isoformat()}와 일치하는 귀국 후보가 없습니다."
        )
        if detected_summaries:
            basis = f"{basis} 후보: {' / '.join(detected_summaries[:4])}"
    else:
        basis = "귀국 후보 결과에서 확정 가능한 단일 일정을 찾지 못했습니다."

    return {
        "baseDepartureDate": base_departure_date,
        "routeKey": route_key,
        "detectedReturnDate": "",
        "detectedReturnTime": "",
        "detectedArrivalDate": "",
        "detectedArrivalTime": "",
        "targetArrivalDate": target_arrival_date.isoformat(),
        "scheduleDays": product_days,
        "nights": "",
        "status": status,
        "basis": basis,
        "candidates": candidate_results,
    }


def load_fare_results(path: str | Path | None) -> dict[tuple[str, str], object]:
    if not path:
        return {}
    result_path = Path(path)
    if not result_path.exists():
        raise FileNotFoundError(f"요금 결과 파일을 찾지 못했습니다: {result_path}")
    payload = json.loads(result_path.read_text(encoding="utf-8"))
    rows = payload.get("rows", payload) if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("요금 결과 JSON은 rows 배열 또는 배열 형식이어야 합니다.")

    fare_cells: dict[tuple[str, str], object] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        route_key = str(row.get("routeKey") or row.get("route_key") or "").strip()
        dep_date = str(row.get("baseDepartureDate") or row.get("date") or row.get("depDate") or "").strip()
        if not route_key or not dep_date:
            continue
        fare_cells[(route_key, dep_date)] = fare_cell_info_from_result_row(row)
    return fare_cells


def fare_cell_info_from_result_row(row: dict[str, object]) -> dict[str, object]:
    status = str(row.get("status") or row.get("fareStatus") or "").strip()
    value = row.get("fare", row.get("roundTripFare", row.get("value")))
    number = to_number(value)
    status_label = fare_status_label(status)
    reason = user_facing_issue_reason(status, str(row.get("reason") or ""))

    if number is not None:
        return {"value": number, "statusLabel": "계산완료", "reason": ""}
    if status_label == "마감":
        return {"value": "마감", "statusLabel": "마감", "reason": reason or "판매 가능한 운임/좌석 없음"}
    if status_label == "계산완료":
        return {"value": "", "statusLabel": "확인필요", "reason": reason or "요금 값이 비어 있습니다."}
    return {"value": status_label, "statusLabel": status_label, "reason": reason}


def fare_status_label(status: str) -> str:
    text = str(status or "").strip().lower()
    if text == "priced":
        return "계산완료"
    if text in {"no_flight", "departure_no_flight", "return_no_flight"}:
        return "비운항"
    if text in {"closed", "sold_out", "마감", "인디비마감"}:
        return "마감"
    if "collect_failed" in text:
        return "조회오류"
    if "raw_missing" in text:
        return "미수집"
    if "command_mismatch" in text:
        return "원문불일치"
    if "schedule_mismatch" in text:
        return "일정불일치"
    if "parse_failed" in text or "ambiguous" in text:
        return "원문확인"
    if text == "fare_route_missing":
        return "운임DB없음"
    if text == "class_missing":
        return "클래스없음"
    if not text:
        return "조회전"
    return "확인필요"


def user_facing_issue_reason(status: str, reason: str) -> str:
    status_text = str(status or "").strip().lower()
    reason_text = str(reason or "").strip()
    lower_reason = reason_text.lower()
    if "stale element reference" in lower_reason:
        return "TOPAS 화면이 갱신되어 조회 결과를 읽는 중 오류가 발생했습니다. 재시도 필요"
    if "timeout" in lower_reason or "timed out" in lower_reason:
        return "TOPAS 응답 대기 시간이 초과되었습니다. 재시도 필요"
    if "collect_failed" in status_text:
        return reason_text or "TOPAS 조회 중 오류가 발생했습니다. 재시도 필요"
    if status_text in {"no_flight", "departure_no_flight", "return_no_flight"}:
        return reason_text or "TOPAS NO FLIGHT: 해당 날짜 비운항 또는 운항편 없음"
    if "raw_missing" in status_text:
        return reason_text or "아직 raw 원문이 수집되지 않았습니다."
    if "command_mismatch" in status_text:
        return reason_text or "요청한 TOPAS 명령과 저장된 원문 명령이 다릅니다. 재수집 필요"
    if "parse_failed" in status_text:
        return reason_text or "TOPAS 원문은 있으나 항공편 라인을 해석하지 못했습니다."
    if "ambiguous" in status_text:
        return reason_text or "확정 가능한 항공편 후보가 2개 이상입니다."
    if status_text == "fare_route_missing":
        return reason_text or "운임 DB에 해당 노선이 없습니다."
    if status_text == "class_missing":
        return reason_text or "TOPAS 원문에서 계산 가능한 클래스 토큰을 찾지 못했습니다."
    return reason_text


def normalize_fare_cell_info(value: object) -> dict[str, object]:
    if isinstance(value, dict):
        display_value = value.get("value", "")
        status_label = str(value.get("statusLabel") or "").strip()
        reason = str(value.get("reason") or "").strip()
        if status_label:
            return {"value": display_value, "statusLabel": status_label, "reason": reason}
    number = to_number(value)
    if number is not None:
        return {"value": number, "statusLabel": "계산완료", "reason": ""}
    text = str(value or "").strip()
    if text:
        return {"value": text, "statusLabel": text, "reason": ""}
    return {"value": "", "statusLabel": "조회전", "reason": "수집/계산 결과 없음"}


def is_numeric_cell_value(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def excel_display_length(value: object) -> int:
    if isinstance(value, datetime):
        return len("yyyy-mm-dd")
    if isinstance(value, date):
        return len("yyyy-mm-dd")
    if is_numeric_cell_value(value):
        number = float(value)
        if number.is_integer():
            return len(f"{int(number):,}")
        return len(f"{number:,.2f}".rstrip("0").rstrip("."))
    return len(str(value or ""))


def summarize_fare_cell_infos(infos: list[dict[str, object]]) -> tuple[str, str]:
    if any(is_numeric_cell_value(info.get("value")) for info in infos):
        problem_infos = [
            info
            for info in infos
            if not is_numeric_cell_value(info.get("value"))
            and str(info.get("statusLabel") or "") not in {"", "마감", "비운항", "조회전"}
        ]
        if problem_infos:
            return "일부확인필요", compact_reason(problem_infos)
        return "계산완료", ""

    for label in ("조회오류", "미수집", "원문불일치", "일정불일치", "원문확인", "운임DB없음", "클래스없음", "확인필요"):
        matches = [info for info in infos if str(info.get("statusLabel") or "") == label]
        if matches:
            return label, compact_reason(matches)
    if any(str(info.get("statusLabel") or "") == "비운항" for info in infos):
        return "비운항", "해당 날짜 TOPAS NO FLIGHT"
    if any(str(info.get("statusLabel") or "") == "마감" for info in infos):
        return "마감", "판매 가능한 운임/좌석 없음"
    return "조회전", "수집/계산 결과 없음"


def compact_reason(infos: list[dict[str, object]], max_length: int = 180) -> str:
    values: list[str] = []
    for info in infos:
        status_label = str(info.get("statusLabel") or "").strip()
        reason = str(info.get("reason") or "").strip()
        text = reason or status_label
        if text and text not in values:
            values.append(text)
    joined = " / ".join(values)
    if len(joined) <= max_length:
        return joined
    return joined[: max_length - 3].rstrip() + "..."


def summary_price_formula(fare_range: str) -> str:
    return f'=IF(COUNT({fare_range})=0,"예약마감",MIN({fare_range}))'


def to_number(value: object) -> int | float | None:
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").replace(",", "").strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return number


def js_round(value: float) -> int:
    return int(math.floor(float(value) + 0.5))


def fare_route_for_master(master: FlightMaster) -> str:
    if master.fare_route:
        return master.fare_route
    mapped = FARE_ROUTE_MAP.get(normalize_fare_route_key(master.raw_key))
    if mapped:
        return mapped
    return f"{master.route}-{master.airline}"


def normalize_fare_route_key(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").upper())


def available_class_tokens(classes: dict[str, str] | None) -> tuple[str, ...]:
    tokens = []
    for code, value in (classes or {}).items():
        text = str(value).strip().upper()
        if text.isdigit() and 4 <= int(text) <= 9:
            tokens.append(f"{str(code).strip().upper()}{text}")
    return tuple(tokens)


def reclass(token: str) -> str:
    return "".join(ch for ch in str(token).upper() if not ch.isdigit())


def load_actual_fare_snapshot(
    cache_path: str | Path | None = None,
    refresh: bool = False,
    prefer_cache_hours: float = 12,
) -> dict[str, object]:
    target_cache = Path(cache_path or LOCAL_FARE_CACHE_PATH)
    if not target_cache.exists() and FARE_SEED_CACHE_PATH.exists():
        target_cache.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(FARE_SEED_CACHE_PATH, target_cache)

    try:
        snapshot = load_fare_snapshot(
            {},
            cache_path=target_cache,
            prefer_cache_within_hours=None if refresh else prefer_cache_hours,
        )
        return {
            "fares": list(snapshot.fares),
            "seasons": list(snapshot.seasons),
            "loadedAt": snapshot.loaded_at,
            "source": snapshot.source,
            "cachePath": str(target_cache),
            "warning": snapshot.warning or "",
        }
    except Exception as exc:
        raise RuntimeError(f"앱 내 운임 DB를 불러오지 못했습니다: {exc}") from exc


def build_fare_map(fares: list[dict[str, object]], route: str) -> dict[str, dict[str, float]]:
    fare_map: dict[str, dict[str, float]] = {}
    for item in fares:
        if str(item.get("route", "")).strip() != route:
            continue
        fare_type = str(item.get("type", "기준")).strip() or "기준"
        class_code = str(item.get("classCode", "")).strip().upper()
        round_trip = to_number(item.get("roundTripFare"))
        if not class_code or round_trip is None:
            continue
        fare_map.setdefault(fare_type, {})[class_code] = float(round_trip) / 2
    return fare_map


def find_season_type(seasons: list[dict[str, object]], route: str, travel_date: str) -> str:
    for item in seasons:
        if str(item.get("route", "")).strip() != route:
            continue
        if str(item.get("startDate", "")) <= travel_date <= str(item.get("endDate", "")):
            return str(item.get("type") or "기준")
    return "기준"


def select_one_way_fare(
    classes: tuple[str, ...],
    season_type: str,
    fare_map: dict[str, dict[str, float]],
) -> dict[str, object]:
    base_map = fare_map.get("기준", {})
    season_map = fare_map.get(season_type, {})
    base_y = base_map.get("Y")
    fallback = base_y * 1.2 if base_y else 0
    lowest = float("inf")
    selected_class = ""
    class_fares: list[dict[str, object]] = []

    for token in classes:
        class_code = reclass(token)
        fare = season_map.get(class_code, base_map.get(class_code))
        class_fares.append(
            {
                "token": token,
                "classCode": class_code,
                "fare": fare,
                "seasonFare": season_map.get(class_code),
                "baseFare": base_map.get(class_code),
            }
        )
        if fare is not None and fare < lowest:
            lowest = fare
            selected_class = class_code

    if lowest == float("inf"):
        return {
            "selectedClass": "N/A",
            "fare": fallback,
            "isClosed": True,
            "classFares": class_fares,
        }
    return {
        "selectedClass": selected_class,
        "fare": lowest,
        "isClosed": False,
        "classFares": class_fares,
    }


def calculate_run_fares(
    log_dir: str | Path,
    cache_path: str | Path | None = None,
    refresh_fare_db: bool = False,
    write_excel: bool = True,
    force_excel_with_pending: bool = False,
) -> Path:
    log_path = Path(log_dir)
    run_json_path = log_path / "run.json"
    if not run_json_path.exists():
        raise FileNotFoundError(f"run.json을 찾지 못했습니다: {run_json_path}")

    if not (log_path / "departure-results.json").exists() or not (log_path / "return-night-results.json").exists():
        process_run_raw_outputs(log_path)

    run_doc = json.loads(run_json_path.read_text(encoding="utf-8"))
    departure_rows = json.loads((log_path / "departure-results.json").read_text(encoding="utf-8")).get("rows", [])
    night_rows = json.loads((log_path / "return-night-results.json").read_text(encoding="utf-8")).get("rows", [])
    masters = {master.key: master for master in load_flight_masters() if master.enabled}
    snapshot = load_actual_fare_snapshot(cache_path=cache_path, refresh=refresh_fare_db)
    fares = [dict(item) for item in snapshot["fares"]]
    seasons = [dict(item) for item in snapshot["seasons"]]
    available_routes = {str(item.get("route", "")) for item in fares if item.get("route")}
    dep_by_key = {
        (str(row.get("routeKey", "")), str(row.get("baseDepartureDate", ""))): row
        for row in departure_rows
        if isinstance(row, dict)
    }

    rows: list[dict[str, object]] = []
    for night_row in night_rows:
        if not isinstance(night_row, dict):
            continue
        route_key = str(night_row.get("routeKey", ""))
        dep_date = str(night_row.get("baseDepartureDate", ""))
        master = masters.get(route_key)
        if master is None:
            continue
        fare_route = fare_route_for_master(master)
        base_row: dict[str, object] = {
            "routeKey": route_key,
            "rawKey": master.raw_key,
            "fareRoute": fare_route,
            "route": master.route,
            "airline": master.airline,
            "depFlight": master.dep_flight,
            "retFlight": master.ret_flight,
            "baseDepartureDate": dep_date,
            "detectedReturnDate": night_row.get("detectedReturnDate", ""),
            "detectedReturnTime": night_row.get("detectedReturnTime", ""),
            "detectedArrivalDate": night_row.get("detectedArrivalDate", ""),
            "detectedArrivalTime": night_row.get("detectedArrivalTime", ""),
            "targetArrivalDate": night_row.get("targetArrivalDate", ""),
            "scheduleDays": night_row.get("scheduleDays", ""),
            "nights": night_row.get("nights", ""),
        }
        dep_result = dep_by_key.get((route_key, dep_date))
        dep_status = str(dep_result.get("status") if isinstance(dep_result, dict) else "raw_missing")
        if not dep_result or dep_status == "raw_missing":
            rows.append({**base_row, "status": "departure_raw_missing", "fare": "", "reason": "출발편 raw 미수집"})
            continue
        if dep_status == "collect_failed":
            rows.append(
                {
                    **base_row,
                    "status": "departure_collect_failed",
                    "fare": "",
                    "reason": str(dep_result.get("error") or "출발편 TOPAS 수집 실패"),
                }
            )
            continue
        if dep_status == "no_flight":
            rows.append({**base_row, "status": "departure_no_flight", "fare": "마감", "reason": "출발편 NO FLIGHT"})
            continue
        if dep_status != "detected":
            dep_reason = str(dep_result.get("error") or "").strip() if isinstance(dep_result, dict) else ""
            rows.append(
                {
                    **base_row,
                    "status": f"departure_{dep_status}",
                    "fare": "",
                    "reason": dep_reason or "출발편 원문/파싱 결과 확인 필요",
                }
            )
            continue

        night_status = str(night_row.get("status") or "return_not_confirmed")
        if night_status != "confirmed":
            status = f"return_{night_status}" if not night_status.startswith("return_") else night_status
            fare_value = "마감" if night_status == "no_flight" else ""
            rows.append(
                {
                    **base_row,
                    "status": status,
                    "fare": fare_value,
                    "reason": str(night_row.get("basis") or "귀국 박수 미확정"),
                }
            )
            continue

        if fare_route not in available_routes:
            rows.append({**base_row, "status": "fare_route_missing", "fare": "", "reason": "운임 DB에 노선이 없습니다."})
            continue

        ret_result = select_confirmed_return_candidate(night_row)
        if not ret_result or ret_result.get("status") != "detected":
            ret_status = str(ret_result.get("status") if isinstance(ret_result, dict) else "raw_missing")
            ret_reason = str(ret_result.get("error") or "").strip() if isinstance(ret_result, dict) else ""
            rows.append(
                {
                    **base_row,
                    "status": f"return_{ret_status}",
                    "fare": "",
                    "reason": ret_reason or "귀국편 원문/파싱 결과 확인 필요",
                }
            )
            continue

        dep_classes = first_match_class_tokens(dep_result)
        ret_classes = first_match_class_tokens(ret_result)
        if not dep_classes or not ret_classes:
            rows.append({**base_row, "status": "class_missing", "fare": "", "reason": "가용 클래스 토큰 없음"})
            continue

        fare_map = build_fare_map(fares, fare_route)
        dep_season = find_season_type(seasons, fare_route, dep_date)
        ret_date = str(night_row.get("detectedReturnDate", ""))
        ret_season = find_season_type(seasons, fare_route, ret_date)
        dep_calc = select_one_way_fare(dep_classes, dep_season, fare_map)
        ret_calc = select_one_way_fare(ret_classes, ret_season, fare_map)
        is_closed = bool(dep_calc["isClosed"] or ret_calc["isClosed"])
        total_fare = float(dep_calc["fare"]) + float(ret_calc["fare"])
        rows.append(
            {
                **base_row,
                "status": "closed" if is_closed else "priced",
                "fare": "마감" if is_closed else js_round(total_fare),
                "roundTripFare": js_round(total_fare),
                "depFare": js_round(float(dep_calc["fare"])),
                "retFare": js_round(float(ret_calc["fare"])),
                "depSelectedClass": dep_calc["selectedClass"],
                "retSelectedClass": ret_calc["selectedClass"],
                "depClasses": list(dep_classes),
                "retClasses": list(ret_classes),
                "depSeason": dep_season,
                "retSeason": ret_season,
                "depRawFile": dep_result.get("rawFile", ""),
                "retRawFile": ret_result.get("rawFile", ""),
                "reason": "운임 계산 완료" if not is_closed else "등록 운임 클래스 없음",
            }
        )

    pending_summary = collect_pending_summary(log_path)
    actionable_pending = int(pending_summary.get("actionablePendingCommands") or 0)
    forced_with_pending = bool(force_excel_with_pending and actionable_pending)

    fare_results_path = log_path / "fare-results.json"
    fare_results_path.write_text(
        json.dumps(
            {
                "generatedAt": now_iso(),
                "forcedExcelWithPending": forced_with_pending,
                "pendingSummary": pending_summary,
                "fareDb": {
                    "source": snapshot["source"],
                    "loadedAt": snapshot["loadedAt"],
                    "cachePath": snapshot["cachePath"],
                    "warning": snapshot["warning"],
                },
                "rows": rows,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    run_doc["fareResults"] = str(fare_results_path)
    run_doc["fareCalculatedAt"] = now_iso()
    calculated_excel: Path | None = None
    can_generate_final_excel = write_excel and (actionable_pending == 0 or force_excel_with_pending)
    if can_generate_final_excel:
        calculated_excel = write_calculated_excel_for_run(log_path, fare_results_path)
        run_doc["calculatedExcel"] = str(calculated_excel)
        run_doc["excelStatus"] = "final_generated_with_pending" if actionable_pending else "final_generated"
        run_doc["excelGeneratedAt"] = now_iso()
        run_doc["forcedExcelWithPending"] = forced_with_pending
        if actionable_pending:
            run_doc["pendingSummary"] = pending_summary
        issue_labels = {"조회오류", "미수집", "원문불일치", "일정불일치", "원문확인", "운임DB없음", "클래스없음", "확인필요"}
        has_errors = any(fare_status_label(str(row.get("status") or "")) in issue_labels for row in rows)
        run_doc["status"] = "completed_with_errors" if has_errors or actionable_pending else "completed"
    else:
        run_doc.pop("calculatedExcel", None)
        run_doc["forcedExcelWithPending"] = False
        run_doc["excelStatus"] = "pending_final_collection"
        run_doc["pendingSummary"] = pending_summary
    atomic_write_text(run_json_path, json.dumps(run_doc, ensure_ascii=False, indent=2) + "\n")
    event_payload = {
        "fareResults": str(fare_results_path),
        "fareRowCount": len(rows),
        "pricedCount": sum(1 for row in rows if row.get("status") == "priced"),
        "closedCount": sum(1 for row in rows if row.get("status") == "closed"),
        "excelGenerated": bool(calculated_excel),
        "actionablePendingCommands": actionable_pending,
        "forcedExcelWithPending": bool(calculated_excel and forced_with_pending),
        "pendingSummary": pending_summary,
    }
    if calculated_excel is not None:
        event_payload["calculatedExcel"] = str(calculated_excel)
    append_run_event(run_doc, "fare_results_calculated", **event_payload)
    return fare_results_path


def select_confirmed_return_candidate(night_row: dict[str, object]) -> dict[str, object] | None:
    target_date = str(night_row.get("detectedReturnDate", ""))
    target_nights = str(night_row.get("nights", ""))
    for item in night_row.get("candidates", []):
        if not isinstance(item, dict):
            continue
        if str(item.get("candidateReturnDate", "")) == target_date and str(item.get("candidateNights", "")) == target_nights:
            return item
    return None


def first_match_class_tokens(row: dict[str, object]) -> tuple[str, ...]:
    matches = row.get("matches", [])
    if not isinstance(matches, list) or not matches:
        return ()
    first = matches[0]
    if not isinstance(first, dict):
        return ()
    classes = first.get("classes", {})
    return available_class_tokens(classes if isinstance(classes, dict) else {})


def write_calculated_excel_for_run(log_dir: str | Path, fare_results_path: str | Path) -> Path:
    log_path = Path(log_dir)
    run_doc = json.loads((log_path / "run.json").read_text(encoding="utf-8"))
    start = parse_iso_date(str(run_doc["startDate"]))
    end = parse_iso_date(str(run_doc["endDate"]))
    selected_routes = set(run_doc.get("selectedRoutes", []))
    masters = [
        master
        for master in load_flight_masters()
        if master.enabled and (not selected_routes or master.route in selected_routes)
    ]
    fare_cells = load_fare_results(fare_results_path)
    run_id = str(run_doc.get("runId") or log_path.name)
    target_override = str(run_doc.get("calculatedExcelTarget") or "").strip()
    target_path = Path(target_override) if target_override else OUTPUT_DIR / f"항공자동조회_MVP_계산결과_{run_id}.xlsx"
    path = write_route_mvp_excel(target_path, masters, start, end, fare_cells)
    return add_fare_summary_sheet(path, fare_results_path)


def add_fare_summary_sheet(path: str | Path, fare_results_path: str | Path) -> Path:
    workbook_path = Path(path)
    payload = json.loads(Path(fare_results_path).read_text(encoding="utf-8"))
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        rows = []
    fare_db = payload.get("fareDb", {}) if isinstance(payload, dict) else {}
    if not isinstance(fare_db, dict):
        fare_db = {}
    pending_summary = payload.get("pendingSummary", {}) if isinstance(payload, dict) else {}
    if not isinstance(pending_summary, dict):
        pending_summary = {}
    forced_with_pending = bool(payload.get("forcedExcelWithPending")) if isinstance(payload, dict) else False

    wb = load_workbook(workbook_path)
    if "요약" in wb.sheetnames:
        wb.remove(wb["요약"])
    ws = wb.create_sheet("요약", 0)
    ws.freeze_panes = "A8"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    issue_fill = PatternFill("solid", fgColor="FCE4D6")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True)

    ws["A1"] = "항공자동조회 계산 요약"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:I1")

    meta_rows = [
        ("결과 생성시각", payload.get("generatedAt", "") if isinstance(payload, dict) else ""),
        ("운임 DB source", fare_db.get("source", "")),
        ("운임 DB loadedAt", fare_db.get("loadedAt", "")),
        ("운임 DB cachePath", fare_db.get("cachePath", "")),
        ("운임 DB warning", fare_db.get("warning", "")),
        ("엑셀 생성 정책", "pending 포함 생성" if forced_with_pending else "pending 0 기준 생성"),
        ("남은 자동 진행대상", pending_summary.get("actionablePendingCommands", 0)),
        ("미수집 / 재시도대상 / retry초과", "{missing} / {retryable} / {exhausted}".format(
            missing=pending_summary.get("missing", 0),
            retryable=int(pending_summary.get("collect_error_retryable") or 0)
            + int(pending_summary.get("empty_or_truncated_retryable") or 0)
            + int(pending_summary.get("command_mismatch_retryable") or 0),
            exhausted=pending_summary.get("retry_exhausted", 0),
        )),
        ("fare-results", str(Path(fare_results_path))),
    ]
    for row_index, (label, value) in enumerate(meta_rows, start=2):
        ws.cell(row=row_index, column=1, value=label).font = header_font
        ws.cell(row=row_index, column=2, value=value)

    status_counts: dict[str, int] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        status = fare_status_label(str(row.get("status") or "unknown"))
        status_counts[status] = status_counts.get(status, 0) + 1

    status_start = len(meta_rows) + 3
    ws.cell(row=status_start, column=1, value="상태").font = header_font
    ws.cell(row=status_start, column=2, value="건수").font = header_font
    ws.cell(row=status_start, column=1).fill = header_fill
    ws.cell(row=status_start, column=2).fill = header_fill
    for offset, (status, count) in enumerate(sorted(status_counts.items()), start=1):
        ws.cell(row=status_start + offset, column=1, value=status)
        ws.cell(row=status_start + offset, column=2, value=count)

    priced_start = status_start + max(len(status_counts), 1) + 3
    priced_headers = ["routeKey", "출발일", "요금", "박수", "귀국일", "귀국시간", "출발클래스", "귀국클래스"]
    ws.cell(row=priced_start, column=1, value="계산 완료").font = header_font
    for column_index, header in enumerate(priced_headers, start=1):
        cell = ws.cell(row=priced_start + 1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill

    priced_row_index = priced_start + 2
    for row in rows:
        if not isinstance(row, dict) or row.get("status") != "priced":
            continue
        values = [
            row.get("routeKey", ""),
            row.get("baseDepartureDate", ""),
            row.get("fare", ""),
            row.get("nights", ""),
            row.get("detectedReturnDate", ""),
            row.get("detectedReturnTime", ""),
            row.get("depSelectedClass", ""),
            row.get("retSelectedClass", ""),
        ]
        for column_index, value in enumerate(values, start=1):
            ws.cell(row=priced_row_index, column=column_index, value=value)
        priced_row_index += 1

    issue_start = max(priced_row_index + 2, priced_start + 5)
    issue_headers = ["routeKey", "출발일", "상태", "상세상태", "사유", "fareRoute", "출발편", "귀국편", "rawKey"]
    ws.cell(row=issue_start, column=1, value="확인 필요").font = header_font
    for column_index, header in enumerate(issue_headers, start=1):
        cell = ws.cell(row=issue_start + 1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = issue_fill

    issue_row_index = issue_start + 2
    for row in rows:
        if not isinstance(row, dict) or row.get("status") == "priced":
            continue
        detail_status = str(row.get("status", ""))
        values = [
            row.get("routeKey", ""),
            row.get("baseDepartureDate", ""),
            fare_status_label(detail_status),
            detail_status,
            user_facing_issue_reason(detail_status, str(row.get("reason") or "")),
            row.get("fareRoute", ""),
            row.get("depFlight", ""),
            row.get("retFlight", ""),
            row.get("rawKey", ""),
        ]
        for column_index, value in enumerate(values, start=1):
            ws.cell(row=issue_row_index, column=column_index, value=value)
        issue_row_index += 1

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 28,
        "B": 16,
        "C": 16,
        "D": 34,
        "E": 20,
        "F": 14,
        "G": 14,
        "H": 18,
        "I": 18,
    }
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width

    return save_workbook_with_fallback(wb, workbook_path)


def process_run_raw_outputs(log_dir: str | Path) -> dict[str, Path]:
    log_path = Path(log_dir)
    run_json_path = log_path / "run.json"
    command_plan_path = log_path / "command-plan.json"
    if not run_json_path.exists():
        raise FileNotFoundError(f"run.json을 찾지 못했습니다: {run_json_path}")
    if not command_plan_path.exists():
        raise FileNotFoundError(f"command-plan.json을 찾지 못했습니다: {command_plan_path}")

    run_doc = json.loads(run_json_path.read_text(encoding="utf-8"))
    command_plan = json.loads(command_plan_path.read_text(encoding="utf-8"))
    masters = {master.key: master for master in load_flight_masters() if master.enabled}
    retry_state = load_retry_state(log_path)
    raw_records = load_raw_store_records(log_path)

    departure_results: list[dict[str, object]] = []
    for command in command_plan.get("departureCommands", []):
        if not isinstance(command, dict):
            continue
        raw_file = Path(str(command.get("rawFile", "")))
        raw_record = raw_record_from_cache(raw_records, command)
        raw_state = raw_response_state_from_record(raw_record, raw_file, command)
        route_key = str(command.get("routeKey", ""))
        master = masters.get(route_key)
        if master is None:
            continue
        result = {
            "routeKey": route_key,
            "baseDepartureDate": command.get("baseDepartureDate", ""),
            "queryDate": command.get("queryDate", ""),
            "command": command.get("command", ""),
            "rawFile": str(raw_file),
            "rawExists": raw_state != "missing",
        }
        raw_text, raw_exists, _read_error, _raw_source = read_raw_response_text_from_record(raw_record, raw_file)
        if raw_exists:
            exhausted_error = retry_exhausted_error_for_raw(retry_state, raw_file)
            if exhausted_error and raw_state in {"collect_error", "empty_or_truncated"}:
                result.update({"status": "collect_failed", "error": exhausted_error, "matches": []})
            elif command_response_mismatch(command, raw_text):
                result.update(
                    {
                        "status": "command_mismatch",
                        "error": command_mismatch_reason(command, raw_text),
                        "matches": [],
                    }
                )
            else:
                parsed = parse_topas_flight_result(
                    raw_text,
                    master.airline,
                    master.dep_flight,
                )
                result.update(parsed)
        else:
            result.update({"status": "raw_missing", "matches": []})
        departure_results.append(result)

    return_candidate_results: list[dict[str, object]] = []
    grouped_return: dict[tuple[str, str], list[dict[str, object]]] = {}
    for command in command_plan.get("returnCandidateCommands", []):
        if not isinstance(command, dict):
            continue
        raw_file = Path(str(command.get("rawFile", "")))
        raw_record = raw_record_from_cache(raw_records, command)
        raw_state = raw_response_state_from_record(raw_record, raw_file, command)
        route_key = str(command.get("routeKey", ""))
        master = masters.get(route_key)
        if master is None:
            continue
        result = {
            "routeKey": route_key,
            "baseDepartureDate": command.get("baseDepartureDate", ""),
            "candidateNights": command.get("candidateNights", ""),
            "candidateReturnDate": command.get("candidateReturnDate", ""),
            "queryDate": command.get("queryDate", ""),
            "command": command.get("command", ""),
            "rawFile": str(raw_file),
            "rawExists": raw_state != "missing",
        }
        raw_text, raw_exists, _read_error, _raw_source = read_raw_response_text_from_record(raw_record, raw_file)
        if raw_exists:
            exhausted_error = retry_exhausted_error_for_raw(retry_state, raw_file)
            if exhausted_error and raw_state in {"collect_error", "empty_or_truncated"}:
                result.update(
                    {
                        "status": "collect_failed",
                        "detectedReturnTime": "",
                        "error": exhausted_error,
                        "matches": [],
                    }
                )
            elif command_response_mismatch(command, raw_text):
                result.update(
                    {
                        "status": "command_mismatch",
                        "detectedReturnTime": "",
                        "error": command_mismatch_reason(command, raw_text),
                        "matches": [],
                    }
                )
            else:
                parsed = parse_return_time_from_topas(
                    raw_text,
                    master.airline,
                    master.ret_flight,
                )
                result.update(parsed)
        else:
            result.update({"status": "raw_missing", "detectedReturnTime": "", "matches": []})
        return_candidate_results.append(result)
        key = (route_key, str(command.get("baseDepartureDate", "")))
        grouped_return.setdefault(key, []).append(result)

    night_results = [
        confirm_return_nights(dep_date, route_key, int(command_plan.get("productDays") or run_doc.get("productDays") or 5), items)
        for (route_key, dep_date), items in sorted(grouped_return.items())
    ]

    departure_path = log_path / "departure-results.json"
    return_candidates_path = log_path / "return-candidate-results.json"
    nights_path = log_path / "return-night-results.json"
    departure_path.write_text(json.dumps({"rows": departure_results}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return_candidates_path.write_text(
        json.dumps({"rows": return_candidate_results}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    nights_path.write_text(json.dumps({"rows": night_results}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    run_doc["departureResults"] = str(departure_path)
    run_doc["returnCandidateResults"] = str(return_candidates_path)
    run_doc["returnNightResults"] = str(nights_path)
    run_doc["processedAt"] = now_iso()
    atomic_write_text(run_json_path, json.dumps(run_doc, ensure_ascii=False, indent=2) + "\n")
    raw_view_outputs = generate_combined_raw_views(log_path)
    append_run_event(
        run_doc,
        "raw_outputs_processed",
        departureResultCount=len(departure_results),
        returnCandidateResultCount=len(return_candidate_results),
        returnNightResultCount=len(night_results),
    )
    outputs = {
        "departureResults": departure_path,
        "returnCandidateResults": return_candidates_path,
        "returnNightResults": nights_path,
    }
    outputs.update(raw_view_outputs)
    return outputs


def write_route_mvp_excel(
    path: Path,
    masters: list[FlightMaster],
    start: date,
    end: date,
    fare_cells: dict[tuple[str, str], object] | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not masters:
        raise ValueError("엑셀로 만들 노선/편명이 없습니다.")

    wb = Workbook()
    wb.remove(wb.active)
    fare_cells = fare_cells or {}

    for route, route_masters in _masters_by_route(masters).items():
        ws = wb.create_sheet(_route_sheet_title(route))
        write_route_sheet(ws, route_masters, start, end, fare_cells)

    return save_workbook_with_fallback(wb, path)


def save_workbook_with_fallback(wb: Workbook, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists() and workbook_path_is_locked(target):
        fallback = unlocked_workbook_path(target)
        wb.save(fallback)
        return fallback
    try:
        wb.save(target)
        return target
    except PermissionError:
        fallback = unlocked_workbook_path(target)
        wb.save(fallback)
        return fallback


def workbook_path_is_locked(path: str | Path) -> bool:
    target = Path(path)
    if not target.exists():
        return False
    try:
        import msvcrt
    except ImportError:
        return False
    try:
        with target.open("r+b") as handle:
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
        return False
    except OSError:
        return True


def unlocked_workbook_path(path: str | Path) -> Path:
    target = Path(path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = target.with_name(f"{target.stem}_재생성_{stamp}{target.suffix}")
    if not base.exists():
        return base
    counter = 2
    while True:
        candidate = target.with_name(f"{target.stem}_재생성_{stamp}_{counter:02d}{target.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _masters_by_route(masters: list[FlightMaster]) -> dict[str, list[FlightMaster]]:
    airline_order = {airline: index for index, airline in enumerate(sorted({master.airline for master in masters}))}
    grouped: dict[str, list[FlightMaster]] = {}
    for master in masters:
        grouped.setdefault(master.route, []).append(master)
    for route_masters in grouped.values():
        route_masters.sort(key=lambda master: airline_order.get(master.airline, len(airline_order)))
    return grouped


def _route_sheet_title(route: str) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", route.upper())
    return compact[:31] or "ROUTE"


def write_route_sheet(
    ws,
    masters: list[FlightMaster],
    start: date,
    end: date,
    fare_cells: dict[tuple[str, str], object] | None = None,
) -> None:
    fare_cells = fare_cells or {}
    route_title = _route_sheet_title(masters[0].route)
    has_title_row = route_title != "ICNDAD"
    header_row = 2 if has_title_row else 1
    data_start_row = header_row + 1
    first_flight_col = 2
    last_flight_col = first_flight_col + len(masters) - 1
    spacer_col = last_flight_col + 1
    summary_date_col = spacer_col + 1
    summary_price_col = summary_date_col + 1
    summary_air_col = summary_price_col + 1
    summary_status_col = summary_air_col + 1
    summary_reason_col = summary_status_col + 1

    title_fill = PatternFill("solid", fgColor="F3F3F3")
    summary_fill = PatternFill("solid", fgColor="C9DAF8")
    issue_fill = PatternFill("solid", fgColor="FCE4D6")
    center = Alignment(horizontal="center", vertical="center")
    flight_column_widths: dict[int, int] = {}

    if has_title_row:
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=summary_reason_col,
        )
        title_cell = ws.cell(1, 1, "인디비")
        title_cell.fill = title_fill
        title_cell.font = Font(bold=True)
        title_cell.alignment = center

    for col in range(1, summary_reason_col + 1):
        ws.cell(header_row, col).alignment = center

    for offset, master in enumerate(masters):
        cell = ws.cell(header_row, first_flight_col + offset, master.dep_flight)
        cell.alignment = center
        cell.font = Font(bold=True)
        flight_column_widths[first_flight_col + offset] = excel_display_length(master.dep_flight)

    for col, value in (
        (summary_date_col, "날짜"),
        (summary_price_col, "최저가"),
        (summary_air_col, "최저가 항공"),
        (summary_status_col, "상태"),
        (summary_reason_col, "확인사항"),
    ):
        cell = ws.cell(header_row, col, value)
        cell.fill = summary_fill
        cell.font = Font(bold=True)
        cell.alignment = center

    for row_index, dep_date in enumerate(iter_dates(start, end), start=data_start_row):
        date_cell = ws.cell(row_index, 1, dep_date)
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.alignment = center

        row_infos: list[dict[str, object]] = []
        for offset, col in enumerate(range(first_flight_col, last_flight_col + 1)):
            master = masters[offset]
            cell_info = normalize_fare_cell_info(fare_cells.get((master.key, dep_date.isoformat()), ""))
            row_infos.append(cell_info)
            value = cell_info.get("value", "")
            fare_cell = ws.cell(row_index, col, value)
            if is_numeric_cell_value(value):
                fare_cell.number_format = "#,##0"
            elif value:
                fare_cell.fill = issue_fill if str(value) not in {"마감"} else title_fill
            fare_cell.alignment = center
            flight_column_widths[col] = max(flight_column_widths.get(col, 0), excel_display_length(value))

        summary_date = ws.cell(row_index, summary_date_col, f"=A{row_index}")
        summary_date.number_format = "yyyy-mm-dd"
        summary_date.alignment = center

        fare_range = f"{get_column_letter(first_flight_col)}{row_index}:{get_column_letter(last_flight_col)}{row_index}"
        header_range = (
            f"${get_column_letter(first_flight_col)}${header_row}:"
            f"${get_column_letter(last_flight_col)}${header_row}"
        )
        price_cell = ws.cell(
            row_index,
            summary_price_col,
            summary_price_formula(fare_range),
        )
        price_cell.number_format = "#,##0"
        price_cell.alignment = center

        price_cell_ref = f"{get_column_letter(summary_price_col)}{row_index}"
        air_cell = ws.cell(
            row_index,
            summary_air_col,
            f'=IF(ISNUMBER({price_cell_ref}),IFERROR(INDEX({header_range}, 1, MATCH({price_cell_ref}, {fare_range}, 0)), ""),"")',
        )
        air_cell.alignment = center

        status_label, reason = summarize_fare_cell_infos(row_infos)
        status_cell = ws.cell(row_index, summary_status_col, status_label)
        status_cell.alignment = center
        if status_label not in {"계산완료", "마감"}:
            status_cell.fill = issue_fill

        reason_cell = ws.cell(row_index, summary_reason_col, reason)
        reason_cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 13
    for col in range(first_flight_col, last_flight_col + 1):
        width = max(12, flight_column_widths.get(col, 0) + 2)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.column_dimensions[get_column_letter(spacer_col)].width = 13
    ws.column_dimensions[get_column_letter(summary_date_col)].width = 13
    ws.column_dimensions[get_column_letter(summary_price_col)].width = 13
    ws.column_dimensions[get_column_letter(summary_air_col)].width = 12
    ws.column_dimensions[get_column_letter(summary_status_col)].width = 13
    ws.column_dimensions[get_column_letter(summary_reason_col)].width = 34


def default_output_path() -> Path:
    return OUTPUT_DIR / f"항공자동조회_MVP_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def find_chrome_executable() -> Path | None:
    candidates: list[Path] = []
    for env_name in ("PROGRAMFILES", "PROGRAMFILES(X86)", "LOCALAPPDATA"):
        base = os.environ.get(env_name)
        if not base:
            continue
        candidates.append(Path(base) / "Google" / "Chrome" / "Application" / "chrome.exe")
    candidates.extend(
        [
            Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
            Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        ]
    )
    for path in candidates:
        if path.exists():
            return path
    return None


def topas_debug_targets(address: str = TOPAS_DEBUG_ADDRESS, timeout: float = 0.8) -> list[dict[str, object]]:
    with urllib.request.urlopen(f"http://{address}/json/list", timeout=timeout) as response:
        data = response.read().decode("utf-8", errors="replace")
    targets = json.loads(data)
    if not isinstance(targets, list):
        return []
    return [target for target in targets if isinstance(target, dict)]


def target_looks_like_topas(target: dict[str, object]) -> bool:
    title = str(target.get("title", "")).lower()
    url = str(target.get("url", "")).lower()
    return "topassellconnect.com" in url or "topas" in url or "topas" in title or "sell connect" in title


def topas_debug_browser_status(address: str = TOPAS_DEBUG_ADDRESS, timeout: float = 0.8) -> dict[str, object]:
    try:
        targets = topas_debug_targets(address, timeout=timeout)
    except Exception as exc:
        return {"running": False, "targets": [], "topasTargets": [], "error": str(exc)}
    topas_targets = [target for target in targets if target_looks_like_topas(target)]
    return {"running": True, "targets": targets, "topasTargets": topas_targets, "error": ""}


def open_topas_tab_in_debug_browser(address: str = TOPAS_DEBUG_ADDRESS, timeout: float = 1.5) -> None:
    encoded_url = urllib.parse.quote(TOPAS_DEBUG_URL, safe="")
    request = urllib.request.Request(f"http://{address}/json/new?{encoded_url}", method="PUT")
    with urllib.request.urlopen(request, timeout=timeout) as response:
        response.read()


def latest_log_dir() -> Path | None:
    if not LOGS_DIR.exists():
        return None
    candidates = [path for path in LOGS_DIR.iterdir() if path.is_dir()]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def create_run_log(
    start: date,
    end: date,
    product_days: int,
    masters: list[FlightMaster],
    output_path: Path,
) -> dict[str, object]:
    run_id = next_run_id()
    log_dir = LOGS_DIR / run_id
    log_dir.mkdir(parents=True, exist_ok=False)
    run_doc: dict[str, object] = {
        "runId": run_id,
        "startedAt": now_iso(),
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "dateCount": date_count(start, end),
        "productDays": product_days,
        "selectedRoutes": sorted({master.route for master in masters}),
        "selectedRouteCount": len({master.route for master in masters}),
        "selectedMasterKeys": sorted(master.key for master in masters),
        "selectedFlightMasters": [
            {
                "key": master.key,
                "route": master.route,
                "airline": master.airline,
                "depFlight": master.dep_flight,
                "retFlight": master.ret_flight,
            }
            for master in masters
        ],
        "flightMasterCount": len(masters),
        "status": "running",
        "outputExcel": str(output_path),
        "logDir": str(log_dir),
    }
    write_run_json(run_doc)
    append_run_event(
        run_doc,
        "run_started",
        message="실행 로그 생성",
        outputExcel=str(output_path),
    )
    return run_doc


def _same_path(left: str | Path, right: str | Path) -> bool:
    try:
        return Path(left).resolve() == Path(right).resolve()
    except Exception:
        return str(left).strip().lower() == str(right).strip().lower()


def find_source_log_for_raw_store(raw_store_path: str | Path) -> Path | None:
    raw_store = Path(raw_store_path)
    candidates: list[Path] = []
    run_id = raw_store.parent.name
    if raw_store.name.lower() == RAW_STORE_FILENAME and run_id:
        if raw_store.parent.parent.name.lower() == "raw":
            output_root = raw_store.parent.parent.parent
            candidates.append(output_root / "logs" / run_id)
        candidates.append(LOGS_DIR / run_id)

    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        run_path = candidate / "run.json"
        if not run_path.exists():
            continue
        try:
            run_doc = json.loads(run_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if _same_path(run_doc.get("rawStore", ""), raw_store) or candidate.name == run_id:
            return candidate

    if LOGS_DIR.exists():
        for candidate in LOGS_DIR.iterdir():
            run_path = candidate / "run.json"
            if not run_path.exists():
                continue
            try:
                run_doc = json.loads(run_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if _same_path(run_doc.get("rawStore", ""), raw_store):
                return candidate
    return None


def load_source_run_for_raw_store(raw_store_path: str | Path) -> tuple[Path, dict[str, object]]:
    raw_store = Path(raw_store_path)
    if not raw_store.exists():
        raise FileNotFoundError(f"raw-store.sqlite 파일을 찾지 못했습니다: {raw_store}")
    source_log_dir = find_source_log_for_raw_store(raw_store)
    if source_log_dir is None:
        raise FileNotFoundError("raw-store.sqlite와 연결된 원본 작업(run.json)을 찾지 못했습니다.")
    run_path = source_log_dir / "run.json"
    run_doc = json.loads(run_path.read_text(encoding="utf-8"))
    return source_log_dir, run_doc


def recalculate_raw_store_excel(
    raw_store_path: str | Path,
    product_days: int,
    force_excel_with_pending: bool = True,
) -> tuple[Path, dict[str, object]]:
    raw_store = Path(raw_store_path)
    if not raw_store.exists():
        raise FileNotFoundError(f"raw-store.sqlite 파일을 찾지 못했습니다: {raw_store}")
    if not (2 <= product_days <= 30):
        raise ValueError("상품일수는 2~30 사이로 입력해 주세요.")

    source_log_dir, source_run_doc = load_source_run_for_raw_store(raw_store)
    start = parse_iso_date(str(source_run_doc["startDate"]))
    end = parse_iso_date(str(source_run_doc["endDate"]))
    selected_routes = set(str(route) for route in source_run_doc.get("selectedRoutes", []) if route)
    masters = [
        master
        for master in load_flight_masters()
        if master.enabled and (not selected_routes or master.route in selected_routes)
    ]
    if not masters:
        raise ValueError("원본 작업에서 재계산할 노선을 찾지 못했습니다.")

    source_run_id = str(source_run_doc.get("runId") or source_log_dir.name)
    run_id_preview = next_run_id()
    target_path = OUTPUT_DIR / f"항공자동조회_RAW재계산_{source_run_id}_{product_days}일_{run_id_preview}.xlsx"
    run_log: dict[str, object] | None = None
    try:
        run_log = create_run_log(start, end, product_days, masters, target_path)
        run_id = str(run_log["runId"])
        target_path = OUTPUT_DIR / f"항공자동조회_RAW재계산_{source_run_id}_{product_days}일_{run_id}.xlsx"
        run_log.update(
            {
                "rawRecalculation": True,
                "sourceRunId": source_run_id,
                "sourceLogDir": str(source_log_dir),
                "sourceRawStore": str(raw_store),
                "calculatedExcelTarget": str(target_path),
            }
        )
        command_plan = build_command_plan(run_id, masters, start, end, product_days)
        raw_policy = command_plan.get("rawPathPolicy", {}) if isinstance(command_plan.get("rawPathPolicy"), dict) else {}
        raw_policy["rawStore"] = str(raw_store)
        raw_policy["sourceRawStore"] = str(raw_store)
        command_plan["rawPathPolicy"] = raw_policy
        write_command_plan(run_log, command_plan)
        append_run_event(
            run_log,
            "raw_store_recalculation_started",
            sourceRunId=source_run_id,
            sourceRawStore=str(raw_store),
            productDays=product_days,
        )
        process_run_raw_outputs(run_log["logDir"])
        calculate_run_fares(run_log["logDir"], force_excel_with_pending=force_excel_with_pending)
        final_run_doc = json.loads((Path(str(run_log["logDir"])) / "run.json").read_text(encoding="utf-8"))
        excel_path = Path(str(final_run_doc.get("calculatedExcel") or target_path))
        return excel_path, final_run_doc
    except Exception as exc:
        if run_log is not None:
            finish_run_log(
                run_log,
                "failed",
                error=str(exc),
                rawRecalculation=True,
                sourceRawStore=str(raw_store),
                productDays=product_days,
            )
        raise


def next_run_id() -> str:
    base = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = base
    suffix = 1
    while (LOGS_DIR / candidate).exists():
        candidate = f"{base}_{suffix:02d}"
        suffix += 1
    return candidate


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def format_duration(seconds: float | None) -> str:
    if seconds is None or not math.isfinite(seconds):
        return "계산 중"
    seconds = max(0, int(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}시간 {minutes:02d}분"
    if minutes:
        return f"{minutes}분 {secs:02d}초"
    return f"{secs}초"


def write_run_json(run_doc: dict[str, object]) -> None:
    log_dir = Path(str(run_doc["logDir"]))
    log_dir.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(run_doc, ensure_ascii=False, indent=2)
    atomic_write_text(log_dir / "run.json", payload + "\n")


def append_run_event(run_doc: dict[str, object], event_type: str, **payload: object) -> None:
    event = {
        "timestamp": now_iso(),
        "event": event_type,
        **payload,
    }
    log_dir = Path(str(run_doc["logDir"]))
    log_dir.mkdir(parents=True, exist_ok=True)
    with (log_dir / "events.jsonl").open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False) + "\n")


def finish_run_log(run_doc: dict[str, object], status: str, **updates: object) -> None:
    run_doc.update(updates)
    run_doc["status"] = status
    if status == "completed":
        run_doc["completedAt"] = now_iso()
        event_type = "run_completed"
    elif status == "failed":
        run_doc["failedAt"] = now_iso()
        event_type = "run_failed"
    else:
        event_type = "run_status_changed"
    write_run_json(run_doc)
    append_run_event(run_doc, event_type, **updates)


class AirAutoLookupApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("항공자동조회 MVP")
        self.root.geometry("1180x760")
        self.root.minsize(1040, 660)

        self.masters = [master for master in load_flight_masters() if master.enabled]
        self.routes = sorted({master.route for master in self.masters})
        today = date.today()
        self.start_date = tk.StringVar(value=today.isoformat())
        self.end_date = tk.StringVar(value=default_end_date(today).isoformat())
        self.product_days = tk.StringVar(value="5")
        self.current_log_dir = tk.StringVar(value="")
        self.collect_timeout = tk.StringVar(value="80")
        self.recalc_raw_store = tk.StringVar(value="")
        self.recalc_product_days = tk.StringVar(value="5")
        self.recalc_source_summary = tk.StringVar(value="raw-store.sqlite 파일을 선택해 주세요.")
        self.recalc_result_path = tk.StringVar(value="")
        self.route_vars: dict[str, tk.BooleanVar] = {}
        self.master_vars: dict[str, tk.BooleanVar] = {}
        self.master_checkbuttons: dict[str, ttk.Checkbutton] = {}
        self.master_group_frames: dict[str, ttk.LabelFrame] = {}
        self.master_detail_visible = tk.BooleanVar(value=False)
        self.master_detail_toggle_text = tk.StringVar(value="펼쳐보기")
        self.master_detail_summary_text = tk.StringVar(value="")
        self.route_checks_frame: ttk.Frame | None = None
        self.master_detail_frame: ttk.Frame | None = None
        self.manifest_items: list[dict[str, object]] = []
        self.manifest_selected_index: int | None = None
        self.manifest_path_text = tk.StringVar(value=str(HOTELS_MANIFEST_PATH))
        self.manifest_summary_text = tk.StringVar(value="")
        self.manifest_search_text = tk.StringVar(value="")
        self.manifest_sort_column = tk.StringVar(value="route")
        self.manifest_sort_label = tk.StringVar(value="노선그룹")
        self.manifest_sort_desc = tk.BooleanVar(value=False)
        self.manifest_enabled = tk.BooleanVar(value=True)
        self.manifest_origin = tk.StringVar(value="ICN")
        self.manifest_destination = tk.StringVar(value="")
        self.manifest_route = tk.StringVar(value="")
        self.manifest_fare_route = tk.StringVar(value="")
        self.manifest_airline = tk.StringVar(value="")
        self.manifest_dep_flight = tk.StringVar(value="")
        self.manifest_ret_flight = tk.StringVar(value="")
        self.manifest_ret_time = tk.StringVar(value="")
        self.manifest_product_days = tk.StringVar(value="5")
        self.action_buttons: list[ttk.Button] = []
        self.primary_run_button: ttk.Button | None = None
        self.resume_run_button: ttk.Button | None = None
        self.excel_open_button: ttk.Button | None = None
        self.stop_button: ttk.Button | None = None
        self.recalc_run_button: ttk.Button | None = None
        self.log: tk.Text | None = None
        self.recalc_log: tk.Text | None = None
        self.manifest_tree: ttk.Treeview | None = None
        self.manifest_log: tk.Text | None = None
        self._command_running = False
        self._background_running = False
        self._background_process: subprocess.Popen[str] | None = None
        self._close_after_background = False
        self.progress_percent = tk.DoubleVar(value=0.0)
        self.progress_text = tk.StringVar(value="진행 대기")
        self.progress_detail = tk.StringVar(value="실행을 시작하면 진행률과 예상 남은 시간이 표시됩니다.")
        self.progress_bar: ttk.Progressbar | None = None
        self._progress_after_id: str | None = None
        self._progress_running = False
        self._progress_title = ""
        self._progress_start_monotonic: float | None = None
        self._progress_start_iso = ""
        self._progress_initial_pending: int | None = None
        self._event_log_after_id: str | None = None
        self._event_log_running = False
        self._event_log_dir: Path | None = None
        self._event_log_offset = 0
        self._event_log_remainder = ""
        self._event_log_last_ac1_seen: dict[str, int] = {}
        self._event_log_direct_collected_count = 0
        self._app_icon_image: tk.PhotoImage | None = None

        self.set_window_icon()
        self._build_ui()
        self.root.title(f"항공자동조회 {APP_VERSION}")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self._refresh_route_summary()
        self._log(f"마스터 {len(self.masters)}개 로드 완료")
        self.log_current_run_summary()
        self.refresh_progress_panel()
        self.root.after(1000, self.start_startup_update_check)

    def _build_ui(self) -> None:
        outer = ttk.Frame(self.root, padding=12)
        outer.pack(fill=tk.BOTH, expand=True)

        notebook = ttk.Notebook(outer)
        notebook.pack(fill=tk.BOTH, expand=True)
        auto_tab = ttk.Frame(notebook, padding=0)
        recalc_tab = ttk.Frame(notebook, padding=0)
        manifest_tab = ttk.Frame(notebook, padding=0)
        notebook.add(auto_tab, text="자동조회")
        notebook.add(recalc_tab, text="RAW 재계산")
        notebook.add(manifest_tab, text="조회 노선 관리")

        control = ttk.LabelFrame(auto_tab, text="조회 조건", padding=10)
        control.pack(fill=tk.X)

        ttk.Label(control, text="조회 시작일").grid(row=0, column=0, sticky="w")
        ttk.Entry(control, textvariable=self.start_date, width=14).grid(row=0, column=1, padx=(6, 16), sticky="w")

        ttk.Label(control, text="조회 종료일").grid(row=0, column=2, sticky="w")
        ttk.Entry(control, textvariable=self.end_date, width=14).grid(row=0, column=3, padx=(6, 16), sticky="w")

        ttk.Label(control, text="상품일수").grid(row=0, column=4, sticky="w")
        ttk.Entry(control, textvariable=self.product_days, width=8).grid(row=0, column=5, padx=(6, 16), sticky="w")

        self.primary_run_button = self._add_action_button(control, "▶ 새 실행", self.run_primary_action, row=0, column=6, padx=(6, 0))
        self.resume_run_button = self._add_action_button(
            control,
            "↻ 이어 실행",
            lambda: self.collect_topas("all"),
            row=0,
            column=7,
            padx=(6, 0),
        )
        self.resume_run_button.configure(state=tk.DISABLED)
        self.stop_button = ttk.Button(control, text="■ 정지", command=self.stop_collection, state=tk.DISABLED)
        self.stop_button.grid(row=0, column=8, padx=(6, 0), sticky="ew")
        self._add_action_button(control, "작업 불러오기", self.choose_log_dir, row=0, column=9, padx=(6, 0))
        self.excel_open_button = self._add_action_button(
            control,
            "엑셀 열기",
            self.open_calculated_excel,
            row=0,
            column=10,
            padx=(6, 0),
        )
        self.excel_open_button.configure(state=tk.DISABLED)
        self._add_action_button(control, "브라우저 실행", self.open_topas_debug_browser, row=0, column=11, padx=(6, 0))
        control.columnconfigure(12, weight=1)

        route_box = ttk.LabelFrame(auto_tab, text="노선 카테고리", padding=8)
        route_box.pack(fill=tk.X, pady=(10, 0))
        route_buttons = ttk.Frame(route_box)
        route_buttons.pack(side=tk.RIGHT, padx=(12, 0))
        ttk.Button(route_buttons, text="노선 전체 선택", command=lambda: self.set_all_routes(True)).pack(side=tk.LEFT)
        ttk.Button(route_buttons, text="노선 선택 해제", command=lambda: self.set_all_routes(False)).pack(side=tk.LEFT, padx=(6, 0))

        self.route_checks_frame = ttk.Frame(route_box)
        self.route_checks_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        master_box = ttk.LabelFrame(auto_tab, text="항공사/편명 세부 선택", padding=8)
        master_box.pack(fill=tk.X, pady=(10, 0))
        master_header = ttk.Frame(master_box)
        master_header.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(
            master_header,
            text="편명 전체 선택",
            command=lambda: self.set_all_masters(True),
        ).pack(side=tk.LEFT)
        ttk.Button(
            master_header,
            text="편명 선택 해제",
            command=lambda: self.set_all_masters(False),
        ).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(master_header, textvariable=self.master_detail_summary_text, foreground="#555555").pack(
            side=tk.LEFT,
            padx=(12, 0),
        )
        ttk.Button(
            master_header,
            textvariable=self.master_detail_toggle_text,
            command=self.toggle_master_detail_visibility,
            width=10,
        ).pack(side=tk.RIGHT)

        master_grid = ttk.Frame(master_box)
        self.master_detail_frame = master_grid
        master_grid.pack(fill=tk.X, expand=True)
        self.rebuild_route_master_widgets()

        progress_frame = ttk.LabelFrame(auto_tab, text="실시간 진행률", padding=8)
        progress_frame.pack(fill=tk.X, pady=(10, 0))
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_percent,
            maximum=100,
            mode="determinate",
        )
        self.progress_bar.pack(fill=tk.X)
        ttk.Label(progress_frame, textvariable=self.progress_text).pack(anchor="w", pady=(6, 0))
        ttk.Label(progress_frame, textvariable=self.progress_detail, foreground="#555555").pack(anchor="w", pady=(2, 0))

        log_frame = ttk.LabelFrame(auto_tab, text="작업 로그", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        self.log = tk.Text(log_frame, height=14, wrap=tk.WORD)
        self.log.pack(fill=tk.BOTH, expand=True)

        self._build_recalc_tab(recalc_tab)
        self._build_manifest_tab(manifest_tab)

    def _build_recalc_tab(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="수집 RAW 재계산", padding=10)
        box.pack(fill=tk.X)

        ttk.Label(box, text="raw-store.sqlite").grid(row=0, column=0, sticky="w")
        ttk.Entry(box, textvariable=self.recalc_raw_store, width=88).grid(row=0, column=1, padx=(6, 6), sticky="ew")
        ttk.Button(box, text="파일 선택", command=self.choose_recalc_raw_store).grid(row=0, column=2, sticky="ew")

        ttk.Label(box, text="상품일수").grid(row=1, column=0, pady=(8, 0), sticky="w")
        ttk.Entry(box, textvariable=self.recalc_product_days, width=8).grid(row=1, column=1, padx=(6, 6), pady=(8, 0), sticky="w")
        self.recalc_run_button = ttk.Button(box, text="엑셀 생성", command=self.run_raw_store_recalculation)
        self.recalc_run_button.grid(row=1, column=2, pady=(8, 0), sticky="ew")

        ttk.Label(box, textvariable=self.recalc_source_summary, foreground="#555555").grid(
            row=2,
            column=0,
            columnspan=3,
            pady=(8, 0),
            sticky="w",
        )
        box.columnconfigure(1, weight=1)

        result_box = ttk.LabelFrame(parent, text="생성 결과", padding=10)
        result_box.pack(fill=tk.X, pady=(10, 0))
        ttk.Entry(result_box, textvariable=self.recalc_result_path, state="readonly").grid(
            row=0,
            column=0,
            padx=(0, 6),
            sticky="ew",
        )
        ttk.Button(result_box, text="엑셀 열기", command=self.open_recalc_result).grid(row=0, column=1, sticky="ew")
        result_box.columnconfigure(0, weight=1)

        log_frame = ttk.LabelFrame(parent, text="작업 로그", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        self.recalc_log = tk.Text(log_frame, height=14, wrap=tk.WORD)
        self.recalc_log.pack(fill=tk.BOTH, expand=True)

    def _build_manifest_tab(self, parent: ttk.Frame) -> None:
        path_box = ttk.LabelFrame(parent, text="조회 노선그룹 파일", padding=10)
        path_box.pack(fill=tk.X)
        ttk.Entry(path_box, textvariable=self.manifest_path_text, state="readonly").grid(
            row=0,
            column=0,
            padx=(0, 6),
            sticky="ew",
        )
        ttk.Button(path_box, text="현재 파일 다시 읽기", command=self.load_manifest_items).grid(row=0, column=1, padx=(0, 6), sticky="ew")
        ttk.Button(path_box, text="파일 불러오기", command=self.import_manifest_file).grid(row=0, column=2, sticky="ew")
        ttk.Label(path_box, textvariable=self.manifest_summary_text, foreground="#555555").grid(
            row=1,
            column=0,
            columnspan=3,
            pady=(8, 0),
            sticky="w",
        )
        path_box.columnconfigure(0, weight=1)

        body = ttk.Frame(parent)
        body.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        list_box = ttk.LabelFrame(body, text="노선그룹 목록", padding=8)
        list_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        list_toolbar = ttk.Frame(list_box)
        list_toolbar.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(list_toolbar, text="새 노선 추가", command=self.new_manifest_item).pack(side=tk.LEFT)
        ttk.Button(list_toolbar, text="선택 노선 복사", command=self.duplicate_manifest_item).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(list_toolbar, text="검색").pack(side=tk.LEFT, padx=(14, 4))
        ttk.Entry(list_toolbar, textvariable=self.manifest_search_text, width=18).pack(side=tk.LEFT)
        ttk.Button(list_toolbar, text="초기화", command=self.clear_manifest_search).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(list_toolbar, text="정렬").pack(side=tk.LEFT, padx=(14, 4))
        sort_combo = ttk.Combobox(
            list_toolbar,
            textvariable=self.manifest_sort_label,
            values=("노선그룹", "항공사", "출발편", "귀국편", "상태", "운임키"),
            state="readonly",
            width=8,
        )
        sort_combo.pack(side=tk.LEFT)
        sort_combo.bind("<<ComboboxSelected>>", lambda _event: self.update_manifest_sort_from_label())
        ttk.Checkbutton(
            list_toolbar,
            text="내림",
            variable=self.manifest_sort_desc,
            command=self.refresh_manifest_tree,
        ).pack(side=tk.LEFT, padx=(6, 0))
        columns = ("enabled", "route", "airline", "dep", "ret", "fare")
        self.manifest_tree = ttk.Treeview(list_box, columns=columns, show="headings", height=14, selectmode="browse")
        headings = {
            "enabled": "상태",
            "route": "노선그룹",
            "airline": "항공사",
            "dep": "출발편",
            "ret": "귀국편",
            "fare": "운임키",
        }
        widths = {"enabled": 54, "route": 100, "airline": 64, "dep": 84, "ret": 84, "fare": 120}
        for column in columns:
            self.manifest_tree.heading(
                column,
                text=headings[column],
                command=lambda column=column: self.set_manifest_sort_column(column),
            )
            self.manifest_tree.column(column, width=widths[column], anchor="center", stretch=column in {"route", "fare"})
        scrollbar = ttk.Scrollbar(list_box, orient=tk.VERTICAL, command=self.manifest_tree.yview)
        self.manifest_tree.configure(yscrollcommand=scrollbar.set)
        self.manifest_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.manifest_tree.bind("<<TreeviewSelect>>", self.on_manifest_tree_select)
        self.manifest_search_text.trace_add("write", lambda *_args: self.refresh_manifest_tree())

        edit_box = ttk.LabelFrame(body, text="노선그룹 편집", padding=10)
        edit_box.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Checkbutton(edit_box, text="사용", variable=self.manifest_enabled).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 6),
        )
        fields = (
            ("출발지", self.manifest_origin, 8),
            ("도착지", self.manifest_destination, 8),
            ("노선그룹", self.manifest_route, 16),
            ("운임키", self.manifest_fare_route, 16),
            ("항공사", self.manifest_airline, 8),
            ("출발편", self.manifest_dep_flight, 12),
            ("귀국편", self.manifest_ret_flight, 12),
        )
        for row_index, (label, variable, width) in enumerate(fields, start=1):
            ttk.Label(edit_box, text=label).grid(row=row_index, column=0, sticky="w", pady=2)
            ttk.Entry(edit_box, textvariable=variable, width=width).grid(row=row_index, column=1, sticky="ew", pady=2)

        button_row = ttk.Frame(edit_box)
        button_row.grid(row=len(fields) + 1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        ttk.Button(button_row, text="삭제", command=self.delete_manifest_item).pack(side=tk.LEFT)
        ttk.Button(button_row, text="저장", command=self.save_manifest_items).pack(side=tk.RIGHT)
        edit_box.columnconfigure(1, weight=1)

        log_frame = ttk.LabelFrame(parent, text="변경 로그", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=False, pady=(10, 0))
        self.manifest_log = tk.Text(log_frame, height=7, wrap=tk.WORD)
        self.manifest_log.pack(fill=tk.BOTH, expand=True)
        self.load_manifest_items()

    def rebuild_route_master_widgets(self) -> None:
        if self.route_checks_frame is None or self.master_detail_frame is None:
            return
        for child in self.route_checks_frame.winfo_children():
            child.destroy()
        for child in self.master_detail_frame.winfo_children():
            child.destroy()
        self.route_vars.clear()
        self.master_vars.clear()
        self.master_checkbuttons.clear()
        self.master_group_frames.clear()
        self.routes = sorted({master.route for master in self.masters})

        for index, route in enumerate(self.routes):
            var = tk.BooleanVar(value=True)
            self.route_vars[route] = var
            checkbox = ttk.Checkbutton(
                self.route_checks_frame,
                text=route,
                variable=var,
                command=lambda route=route: self.on_route_toggle(route),
            )
            checkbox.grid(row=index // 6, column=index % 6, padx=(0, 14), pady=2, sticky="w")

        for column in range(4):
            self.master_detail_frame.columnconfigure(column, weight=1, uniform="master_route")
        masters_by_route: dict[str, list[FlightMaster]] = {route: [] for route in self.routes}
        for master in self.masters:
            masters_by_route.setdefault(master.route, []).append(master)

        for index, route in enumerate(self.routes):
            group = ttk.LabelFrame(self.master_detail_frame, text=route, padding=(8, 4))
            group.grid(
                row=index // 4,
                column=index % 4,
                padx=(0, 8),
                pady=(0, 8),
                sticky="nsew",
            )
            group.columnconfigure(0, weight=1)
            group.columnconfigure(1, weight=1)
            self.master_group_frames[route] = group
            route_toolbar = ttk.Frame(group)
            route_toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 3))
            ttk.Button(
                route_toolbar,
                text="전체",
                width=5,
                command=lambda route=route: self.set_route_masters(route, True),
            ).pack(side=tk.LEFT)
            ttk.Button(
                route_toolbar,
                text="해제",
                width=5,
                command=lambda route=route: self.set_route_masters(route, False),
            ).pack(side=tk.LEFT, padx=(4, 0))
            for item_index, master in enumerate(masters_by_route.get(route, [])):
                var = tk.BooleanVar(value=True)
                self.master_vars[master.key] = var
                checkbox = ttk.Checkbutton(
                    group,
                    text=self.master_detail_label(master),
                    variable=var,
                    command=lambda master=master: self.on_master_toggle(master),
                )
                checkbox.grid(row=1 + item_index // 2, column=item_index % 2, padx=(0, 10), pady=1, sticky="w")
                self.master_checkbuttons[master.key] = checkbox
        self.update_master_detail_states()
        self.update_master_detail_visibility()

    def load_manifest_items(self) -> None:
        try:
            path = ensure_hotels_manifest()
            self.manifest_path_text.set(str(path))
            self.manifest_items = load_flight_master_items(path)
        except Exception as exc:
            messagebox.showerror("조회 노선그룹 파일 확인", str(exc))
            self._log_manifest(f"조회 노선그룹 파일 로드 실패: {exc}")
            return
        self.manifest_selected_index = None
        self.refresh_manifest_tree()
        self.new_manifest_item(clear_selection=False)
        self._log_manifest(f"조회 노선 {len(self.manifest_items)}개 로드: {HOTELS_MANIFEST_PATH}")

    def refresh_manifest_tree(self) -> None:
        if self.manifest_tree is None:
            return
        for row_id in self.manifest_tree.get_children():
            self.manifest_tree.delete(row_id)
        enabled_count = 0
        route_names: set[str] = set()
        search_text = self.manifest_search_text.get().strip().upper()
        rows: list[tuple[int, FlightMaster, str]] = []
        for index, item in enumerate(self.manifest_items):
            master = flight_master_from_manifest_item(item)
            fare_route = fare_route_for_master(master)
            if master.enabled:
                enabled_count += 1
                route_names.add(master.route)
            searchable = " ".join(
                (
                    master.route,
                    master.airline,
                    master.dep_flight,
                    master.ret_flight,
                    master.raw_key,
                    fare_route,
                    "사용" if master.enabled else "중지",
                )
            ).upper()
            if search_text and search_text not in searchable:
                continue
            rows.append((index, master, fare_route))

        rows.sort(
            key=lambda row: self.manifest_sort_value(row[1], row[2]),
            reverse=self.manifest_sort_desc.get(),
        )
        for index, master, fare_route in rows:
            self.manifest_tree.insert(
                "",
                tk.END,
                iid=str(index),
                values=(
                    "사용" if master.enabled else "중지",
                    master.route,
                    master.airline,
                    master.dep_flight,
                    master.ret_flight,
                    fare_route,
                ),
            )
        summary = f"전체 {len(self.manifest_items)}개 | 사용 {enabled_count}개 | 노선그룹 {len(route_names)}개"
        if search_text:
            summary += f" | 검색 표시 {len(rows)}개"
        self.manifest_summary_text.set(summary)

    def clear_manifest_search(self) -> None:
        self.manifest_search_text.set("")

    def manifest_sort_label_map(self) -> dict[str, str]:
        return {
            "enabled": "상태",
            "route": "노선그룹",
            "airline": "항공사",
            "dep": "출발편",
            "ret": "귀국편",
            "fare": "운임키",
        }

    def update_manifest_sort_from_label(self) -> None:
        labels = {label: column for column, label in self.manifest_sort_label_map().items()}
        self.manifest_sort_column.set(labels.get(self.manifest_sort_label.get(), "route"))
        self.refresh_manifest_tree()

    def set_manifest_sort_column(self, column: str) -> None:
        if self.manifest_sort_column.get() == column:
            self.manifest_sort_desc.set(not self.manifest_sort_desc.get())
        else:
            self.manifest_sort_column.set(column)
            self.manifest_sort_desc.set(False)
        self.manifest_sort_label.set(self.manifest_sort_label_map().get(column, "노선그룹"))
        self.refresh_manifest_tree()

    def manifest_sort_value(self, master: FlightMaster, fare_route: str) -> tuple[object, ...]:
        column = self.manifest_sort_column.get()
        if column == "enabled":
            return (0 if master.enabled else 1, master.route, master.airline, master.dep_flight)
        if column == "airline":
            return (master.airline, master.route, master.dep_flight, master.ret_flight)
        if column == "dep":
            return (master.dep_flight, master.route, master.airline, master.ret_flight)
        if column == "ret":
            return (master.ret_flight, master.route, master.airline, master.dep_flight)
        if column == "fare":
            return (fare_route, master.route, master.airline, master.dep_flight)
        return (master.route, master.airline, master.dep_flight, master.ret_flight)

    def new_manifest_item(self, clear_selection: bool = True) -> None:
        self.manifest_selected_index = None
        self.manifest_enabled.set(True)
        self.manifest_origin.set("ICN")
        self.manifest_destination.set("")
        self.manifest_route.set("")
        self.manifest_fare_route.set("")
        self.manifest_airline.set("")
        self.manifest_dep_flight.set("")
        self.manifest_ret_flight.set("")
        self.manifest_ret_time.set("")
        self.manifest_product_days.set("5")
        if clear_selection and self.manifest_tree is not None:
            self.manifest_tree.selection_remove(self.manifest_tree.selection())

    def duplicate_manifest_item(self) -> None:
        if self.manifest_selected_index is None:
            messagebox.showinfo("노선 복사", "복제할 노선을 먼저 선택해 주세요.")
            return
        self.manifest_selected_index = None
        if self.manifest_tree is not None:
            self.manifest_tree.selection_remove(self.manifest_tree.selection())
        self._log_manifest("선택 노선을 복사 모드로 전환했습니다.")

    def on_manifest_tree_select(self, _event: object | None = None) -> None:
        if self.manifest_tree is None:
            return
        selection = self.manifest_tree.selection()
        if not selection:
            return
        try:
            index = int(selection[0])
            item = self.manifest_items[index]
            master = flight_master_from_manifest_item(item)
        except Exception as exc:
            self._log_manifest(f"선택 항목 로드 실패: {exc}")
            return
        self.manifest_selected_index = index
        self.manifest_enabled.set(master.enabled)
        self.manifest_origin.set(master.origin)
        self.manifest_destination.set(master.destination)
        self.manifest_route.set(master.route)
        self.manifest_fare_route.set(fare_route_for_master(master))
        self.manifest_airline.set(master.airline)
        self.manifest_dep_flight.set(master.dep_flight)
        self.manifest_ret_flight.set(master.ret_flight)
        self.manifest_ret_time.set(master.ret_departure_time or "")
        self.manifest_product_days.set(str(master.default_product_days))

    def manifest_form_item(self) -> dict[str, object]:
        origin = normalize_airport_code(self.manifest_origin.get())
        destination = normalize_airport_code(self.manifest_destination.get())
        airline = normalize_airline_code(self.manifest_airline.get())
        dep_flight = normalize_manifest_flight(airline, self.manifest_dep_flight.get())
        ret_flight = normalize_manifest_flight(airline, self.manifest_ret_flight.get())
        route = self.manifest_route.get().strip().upper() or f"{origin}-{destination}"
        product_days = int(self.manifest_product_days.get())
        item: dict[str, object] = {
            "origin": origin,
            "destination": destination,
            "route": route,
            "fareRoute": self.manifest_fare_route.get().strip().upper(),
            "airline": airline,
            "depFlight": dep_flight,
            "retFlight": ret_flight,
            "retDepartureTime": self.manifest_ret_time.get().strip() or None,
            "defaultProductDays": product_days,
            "enabled": self.manifest_enabled.get(),
        }
        if self.manifest_selected_index is not None:
            old = normalize_flight_master_item(self.manifest_items[self.manifest_selected_index])
            core_keys = ("origin", "destination", "airline", "depFlight", "retFlight")
            if all(str(old[key]) == str(item[key]) for key in core_keys):
                item["key"] = old["key"]
                item["rawKey"] = old["rawKey"]
        return normalize_flight_master_item(item)

    def manifest_form_has_user_input(self) -> bool:
        return any(
            variable.get().strip()
            for variable in (
                self.manifest_destination,
                self.manifest_route,
                self.manifest_fare_route,
                self.manifest_airline,
                self.manifest_dep_flight,
                self.manifest_ret_flight,
            )
        )

    def apply_manifest_form(self) -> bool:
        try:
            item = self.manifest_form_item()
            key = str(item["key"])
            for index, existing in enumerate(self.manifest_items):
                if self.manifest_selected_index is not None and index == self.manifest_selected_index:
                    continue
                if str(existing.get("key") or "").upper() == key:
                    raise ValueError(f"이미 같은 내부키가 있습니다: {key}")
        except Exception as exc:
            messagebox.showerror("노선 입력 확인", str(exc))
            return False

        if self.manifest_selected_index is None:
            self.manifest_items.append(item)
            self.manifest_selected_index = len(self.manifest_items) - 1
            self._log_manifest(f"노선 추가: {item['route']} {item['airline']} {item['depFlight']}/{item['retFlight']}")
        else:
            self.manifest_items[self.manifest_selected_index] = item
            self._log_manifest(f"노선 수정: {item['route']} {item['airline']} {item['depFlight']}/{item['retFlight']}")
        self.refresh_manifest_tree()
        if self.manifest_tree is not None and self.manifest_selected_index is not None:
            row_id = str(self.manifest_selected_index)
            if row_id in self.manifest_tree.get_children():
                self.manifest_tree.selection_set(row_id)
                self.manifest_tree.focus(row_id)
                self.manifest_tree.see(row_id)
        return True

    def delete_manifest_item(self) -> None:
        if self.manifest_selected_index is None:
            messagebox.showinfo("노선 삭제", "삭제할 노선을 먼저 선택해 주세요.")
            return
        item = self.manifest_items[self.manifest_selected_index]
        label = f"{item.get('route')} {item.get('airline')} {item.get('depFlight')}/{item.get('retFlight')}"
        if not messagebox.askyesno("노선 삭제", f"삭제할까요?\n{label}"):
            return
        del self.manifest_items[self.manifest_selected_index]
        self._log_manifest(f"노선 삭제: {label}")
        self.manifest_selected_index = None
        self.refresh_manifest_tree()
        self.new_manifest_item(clear_selection=False)
        self.save_manifest_items()

    def save_manifest_items(self) -> None:
        if self.manifest_selected_index is not None or self.manifest_form_has_user_input():
            if not self.apply_manifest_form():
                return
        try:
            write_hotels_manifest(self.manifest_items)
            self.manifest_items = load_flight_master_items(HOTELS_MANIFEST_PATH)
            self.manifest_path_text.set(str(HOTELS_MANIFEST_PATH))
            self.refresh_manifest_tree()
            self.reload_auto_masters_from_manifest()
        except Exception as exc:
            messagebox.showerror("조회 노선 저장 실패", str(exc))
            self._log_manifest(f"저장 실패: {exc}")
            return
        self._log_manifest(f"저장 완료: {HOTELS_MANIFEST_PATH}")
        messagebox.showinfo("조회 노선 저장", "저장했고 자동조회 탭에 반영했습니다.")

    def reload_auto_masters_from_manifest(self) -> None:
        self.masters = [master for master in load_flight_masters() if master.enabled]
        self.rebuild_route_master_widgets()
        self._refresh_route_summary()
        self._log(f"조회 노선 갱신: {len(self.routes)}개 노선, {len(self.masters)}개 편명")

    def import_manifest_file(self) -> None:
        try:
            selected_path = filedialog.askopenfilename(
                title="조회 노선그룹 파일 불러오기",
                initialdir=str(HOTELS_MANIFEST_PATH.parent),
                filetypes=(("JSON 파일", "*.json"), ("모든 파일", "*.*")),
            )
            if not selected_path:
                return
            path = Path(selected_path)
            items = load_flight_master_items(path)
        except Exception as exc:
            messagebox.showerror("조회 노선그룹 파일 불러오기 실패", str(exc))
            self._log_manifest(f"파일 불러오기 실패: {exc}")
            return

        self.manifest_items = items
        self.manifest_selected_index = None
        self.manifest_path_text.set(str(path))
        self.refresh_manifest_tree()
        self.new_manifest_item(clear_selection=False)
        self._log_manifest(f"파일 불러오기 완료: {path} ({len(items)}개)")
        messagebox.showinfo(
            "파일 불러오기 완료",
            f"노선그룹 {len(items)}개를 불러왔습니다.\n현재 앱에 적용하려면 [저장]을 눌러 주세요.",
        )

    def set_window_icon(self) -> None:
        try:
            if APP_ICON_ICO_PATH.exists():
                self.root.iconbitmap(default=str(APP_ICON_ICO_PATH))
            if APP_ICON_PNG_PATH.exists():
                self._app_icon_image = tk.PhotoImage(file=str(APP_ICON_PNG_PATH))
                self.root.iconphoto(True, self._app_icon_image)
        except Exception as exc:
            # Icon loading should never block the operational GUI.
            print(f"아이콘 로드 실패: {exc}", file=sys.stderr)

    def toggle_master_detail_visibility(self) -> None:
        self.master_detail_visible.set(not self.master_detail_visible.get())
        self.update_master_detail_visibility()

    def update_master_detail_visibility(self) -> None:
        if self.master_detail_frame is None:
            return
        if self.master_detail_visible.get():
            if self.master_detail_frame.winfo_manager() != "pack":
                self.master_detail_frame.pack(fill=tk.X, expand=True)
            self.master_detail_toggle_text.set("접어보기")
        else:
            if self.master_detail_frame.winfo_manager():
                self.master_detail_frame.pack_forget()
            self.master_detail_toggle_text.set("펼쳐보기")

    @staticmethod
    def master_detail_label(master: FlightMaster) -> str:
        dep_flight = master.dep_flight
        ret_flight = master.ret_flight
        if dep_flight.startswith(master.airline):
            dep_flight = dep_flight[len(master.airline) :]
        if ret_flight.startswith(master.airline):
            ret_flight = ret_flight[len(master.airline) :]
        return f"{master.airline} {dep_flight}/{ret_flight}"

    def _refresh_route_summary(self) -> None:
        selected_masters = self.selected_masters()
        selected_route_count = len({master.route for master in selected_masters})
        self.master_detail_summary_text.set(f"선택 {selected_route_count}개 노선 / {len(selected_masters)}개 편명")
        self._log(f"조회 대상: {selected_route_count}개 노선, {len(selected_masters)}개 편명")

    def selected_routes(self) -> set[str]:
        return {route for route, var in self.route_vars.items() if var.get()}

    def on_route_toggle(self, route: str) -> None:
        selected = bool(self.route_vars.get(route, tk.BooleanVar(value=False)).get())
        for master in self.masters:
            if master.route == route and master.key in self.master_vars:
                self.master_vars[master.key].set(selected)
        self.update_master_detail_states()
        self._refresh_route_summary()

    def on_master_toggle(self, master: FlightMaster) -> None:
        if self.master_vars.get(master.key) is not None and self.master_vars[master.key].get():
            route_var = self.route_vars.get(master.route)
            if route_var is not None and not route_var.get():
                route_var.set(True)
        self.update_master_detail_states()
        self._refresh_route_summary()

    def update_master_detail_states(self) -> None:
        for master in self.masters:
            checkbox = self.master_checkbuttons.get(master.key)
            if checkbox is None:
                continue
            route_selected = bool(self.route_vars.get(master.route, tk.BooleanVar(value=False)).get())
            checkbox.configure(state=tk.NORMAL if route_selected else tk.DISABLED)

    def set_all_routes(self, selected: bool) -> None:
        for var in self.route_vars.values():
            var.set(selected)
        for var in self.master_vars.values():
            var.set(selected)
        self.update_master_detail_states()
        self._refresh_route_summary()

    def set_all_masters(self, selected: bool) -> None:
        if selected:
            for var in self.route_vars.values():
                var.set(True)
        for var in self.master_vars.values():
            var.set(selected)
        self.update_master_detail_states()
        self._refresh_route_summary()

    def set_route_masters(self, route: str, selected: bool) -> None:
        route_var = self.route_vars.get(route)
        if route_var is not None and selected:
            route_var.set(True)
        for master in self.masters:
            if master.route == route and master.key in self.master_vars:
                self.master_vars[master.key].set(selected)
        self.update_master_detail_states()
        self._refresh_route_summary()

    def selected_masters(self) -> list[FlightMaster]:
        selected_routes = self.selected_routes()
        if not self.master_vars:
            return [master for master in self.masters if master.route in selected_routes]
        selected: list[FlightMaster] = []
        for master in self.masters:
            var = self.master_vars.get(master.key)
            if master.route in selected_routes and var is not None and var.get():
                selected.append(master)
        return selected

    def _add_action_button(
        self,
        parent,
        text: str,
        command,
        row: int,
        column: int,
        padx=(0, 0),
        pady=(0, 0),
    ) -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command)
        button.grid(row=row, column=column, padx=padx, pady=pady, sticky="ew")
        self.action_buttons.append(button)
        return button

    def has_selected_run_log(self) -> bool:
        value = self.current_log_dir.get().strip()
        return bool(value and (Path(value) / "run.json").exists())

    def has_calculated_excel(self) -> bool:
        if not self.has_selected_run_log():
            return False
        try:
            run_doc = self.read_current_run_doc()
            path_text = str(run_doc.get("calculatedExcel") or "").strip()
            return bool(path_text and Path(path_text).is_file())
        except Exception:
            return False

    def update_primary_run_button(self) -> None:
        if self.primary_run_button is None:
            return
        if self._background_running:
            try:
                pause_requested = self.control_file(PAUSE_CONTROL_FILE).exists()
                stop_requested = self.control_file(STOP_CONTROL_FILE).exists()
            except Exception:
                pause_requested = False
                stop_requested = False
            if stop_requested:
                self.primary_run_button.configure(text="■ 정지 중", state=tk.DISABLED)
            elif pause_requested:
                self.primary_run_button.configure(text="▶ 재개", state=tk.NORMAL)
            else:
                self.primary_run_button.configure(text="⏸ 일시중지", state=tk.NORMAL)
            return
        self.primary_run_button.configure(text="▶ 새 실행")
        if self.resume_run_button is not None:
            self.resume_run_button.configure(state=tk.NORMAL if self.has_selected_run_log() else tk.DISABLED)
        if self.excel_open_button is not None:
            self.excel_open_button.configure(state=tk.NORMAL if self.has_calculated_excel() else tk.DISABLED)

    
    
    def start_startup_update_check(self) -> None:
        def background_check() -> None:
            try:
                import update_client

                manifest = update_client.fetch_available_update(current_version=APP_VERSION, timeout=5.0)
                if manifest and not self._command_running:
                    self.root.after(0, lambda: self.prompt_auto_update(manifest))
            except Exception:
                pass

        threading.Thread(target=background_check, daemon=True).start()

    def prompt_auto_update(self, manifest: dict[str, str]) -> None:
        try:
            import update_client

            latest_ver = manifest.get("version", "최신 버전")
            notes = manifest.get("release_notes") or "새 버전이 출시되었습니다."
            msg = f"새로운 버전({latest_ver})이 출시되었습니다.\n\n[패치 노트]\n{notes}\n\n지금 업데이트를 다운로드하고 설치할까요?"

            if messagebox.askyesno("자동 업데이트 알림", msg):
                self._log(f"새 버전 {latest_ver} 자동 업데이트 진행 중...")
                installer_path = update_client.download_installer(manifest)
                self._log(f"다운로드 완료: {installer_path}")
                update_client.run_installer_and_exit(installer_path)
                self.root.destroy()
        except Exception as exc:
            messagebox.showerror("자동 업데이트 실패", str(exc))
            self._log(f"자동 업데이트 실패: {exc}")

    def check_app_update(self) -> None:
        try:
            import update_client

            self._log(f"현재 버전: {APP_VERSION} - 최신 버전 확인 중...")
            manifest = update_client.fetch_available_update(current_version=APP_VERSION)
            if not manifest:
                messagebox.showinfo("업데이트 확인", f"현재 최신 버전({APP_VERSION})을 사용 중입니다.")
                self._log("최신 버전 사용 중입니다.")
                return

            latest_ver = manifest.get("version", "최신 버전")
            notes = manifest.get("release_notes") or "새 버전이 출시되었습니다."
            msg = f"새로운 버전({latest_ver})이 있습니다.\n\n[패치 노트]\n{notes}\n\n지금 업데이트를 다운로드하고 설치할까요?"

            if messagebox.askyesno("업데이트 알림", msg):
                self._log(f"새 버전 {latest_ver} 다운로드 중...")
                installer_path = update_client.download_installer(manifest)
                self._log(f"다운로드 완료: {installer_path}")
                update_client.run_installer_and_exit(installer_path)
                self.root.destroy()
        except Exception as exc:
            messagebox.showerror("업데이트 확인 실패", str(exc))
            self._log(f"업데이트 확인 실패: {exc}")

    def run_primary_action(self) -> None:
        if self._background_running:
            self.toggle_pause_collection()
            return
        self.auto_run_all()

    def choose_log_dir(self) -> None:
        initial = self.current_log_dir.get() or str(LOGS_DIR)
        selected = filedialog.askdirectory(initialdir=initial)
        if selected:
            self.current_log_dir.set(selected)
            self._log(f"작업 불러오기: {selected}")
            self.load_run_inputs(Path(selected))
            self.update_primary_run_button()
            self.log_current_run_summary()
            self.refresh_progress_panel()

    def choose_recalc_raw_store(self) -> None:
        current = self.recalc_raw_store.get().strip()
        initial_dir = str(Path(current).parent) if current else str(RAW_DIR)
        selected = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="raw-store.sqlite 선택",
            filetypes=[
                ("TOPAS RAW SQLite", "raw-store.sqlite"),
                ("SQLite", "*.sqlite *.db"),
                ("모든 파일", "*.*"),
            ],
        )
        if not selected:
            return
        self.recalc_raw_store.set(selected)
        self.refresh_recalc_raw_store_summary()
        self._log_recalc(f"raw-store 선택: {selected}")

    def refresh_recalc_raw_store_summary(self) -> None:
        raw_store = self.recalc_raw_store.get().strip()
        if not raw_store:
            self.recalc_source_summary.set("raw-store.sqlite 파일을 선택해 주세요.")
            return
        try:
            source_log_dir, run_doc = load_source_run_for_raw_store(raw_store)
            product_days = str(run_doc.get("productDays") or "").strip()
            if product_days:
                self.recalc_product_days.set(product_days)
            self.recalc_source_summary.set(
                "원본 작업 {run_id} | {start}~{end} | 노선 {routes}개 | 원본 상품일수 {days} | {log_dir}".format(
                    run_id=run_doc.get("runId", source_log_dir.name),
                    start=run_doc.get("startDate", ""),
                    end=run_doc.get("endDate", ""),
                    routes=run_doc.get("selectedRouteCount", ""),
                    days=run_doc.get("productDays", ""),
                    log_dir=source_log_dir,
                )
            )
        except Exception as exc:
            self.recalc_source_summary.set(f"원본 작업 확인 실패: {exc}")

    def run_raw_store_recalculation(self) -> None:
        if self._command_running:
            messagebox.showinfo("작업 실행 중", "현재 실행 중인 작업이 끝난 뒤 다시 시도해 주세요.")
            return
        raw_store = self.recalc_raw_store.get().strip()
        try:
            if not raw_store:
                raise FileNotFoundError("raw-store.sqlite 파일을 먼저 선택해 주세요.")
            product_days = int(self.recalc_product_days.get())
            if not (2 <= product_days <= 30):
                raise ValueError("상품일수는 2~30 사이로 입력해 주세요.")
        except Exception as exc:
            messagebox.showerror("RAW 재계산 조건 확인", str(exc))
            return

        self._command_running = True
        self.set_action_buttons_state(tk.DISABLED)
        if self.recalc_run_button is not None:
            self.recalc_run_button.configure(state=tk.DISABLED)
        self.recalc_result_path.set("계산 중...")
        self._log_recalc(f"RAW 재계산 시작: 상품일수 {product_days} | {raw_store}")

        def worker() -> None:
            try:
                excel_path, run_doc = recalculate_raw_store_excel(raw_store, product_days)
            except Exception as exc:
                self.root.after(0, lambda error=exc: self.finish_raw_store_recalculation(None, None, error))
                return
            self.root.after(0, lambda: self.finish_raw_store_recalculation(excel_path, run_doc, None))

        threading.Thread(target=worker, daemon=True).start()

    def finish_raw_store_recalculation(
        self,
        excel_path: Path | None,
        run_doc: dict[str, object] | None,
        error: Exception | None,
    ) -> None:
        self._command_running = False
        self.set_action_buttons_state(tk.NORMAL)
        self.update_primary_run_button()
        if self.recalc_run_button is not None:
            self.recalc_run_button.configure(state=tk.NORMAL)
        if error is not None:
            self.recalc_result_path.set("")
            self._log_recalc(f"RAW 재계산 실패: {error}")
            messagebox.showerror("RAW 재계산 실패", str(error))
            return
        if excel_path is None or run_doc is None:
            self.recalc_result_path.set("")
            self._log_recalc("RAW 재계산 실패: 결과 파일을 확인하지 못했습니다.")
            return
        self.recalc_result_path.set(str(excel_path))
        log_dir = Path(str(run_doc.get("logDir", "")))
        self.log_run_summary(log_dir, target="recalc")
        pending = run_doc.get("pendingSummary", {}) if isinstance(run_doc, dict) else {}
        pending_count = pending.get("actionablePendingCommands", 0) if isinstance(pending, dict) else 0
        self._log_recalc(f"RAW 재계산 엑셀 생성 완료: {excel_path}")
        self._log_recalc(f"남은 자동 진행대상: {pending_count}")
        messagebox.showinfo("RAW 재계산 완료", f"생성 완료:\n{excel_path}")

    def open_recalc_result(self) -> None:
        path_text = self.recalc_result_path.get().strip()
        if not path_text or path_text == "계산 중...":
            messagebox.showinfo("엑셀 열기", "생성된 엑셀 파일이 없습니다.")
            return
        try:
            os.startfile(Path(path_text))
        except Exception as exc:
            messagebox.showerror("엑셀 열기 실패", str(exc))

    def open_topas_debug_browser(self) -> None:
        try:
            status = topas_debug_browser_status(TOPAS_DEBUG_ADDRESS)
            if status["running"]:
                if status["topasTargets"]:
                    self._log(f"TOPAS 디버깅 브라우저 이미 준비됨: {TOPAS_DEBUG_ADDRESS}")
                    return
                try:
                    open_topas_tab_in_debug_browser(TOPAS_DEBUG_ADDRESS)
                    self._log(f"기존 디버깅 브라우저에 TOPAS 탭 열기: {TOPAS_DEBUG_ADDRESS}")
                    return
                except Exception as exc:
                    self._log(f"기존 디버깅 브라우저 TOPAS 탭 열기 실패: {exc}")
                    messagebox.showwarning(
                        "브라우저 확인",
                        "디버깅 Chrome은 이미 켜져 있지만 TOPAS 탭을 자동으로 열지 못했습니다.\n"
                        "열려 있는 디버깅 Chrome에서 TOPAS를 직접 열고 로그인한 뒤 실행해 주세요.",
                    )
                    return

            chrome = find_chrome_executable()
            if chrome is None:
                raise FileNotFoundError("Chrome 실행 파일을 찾지 못했습니다.")
            TOPAS_DEBUG_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
            command = [
                str(chrome),
                f"--remote-debugging-port={TOPAS_DEBUG_PORT}",
                "--remote-debugging-address=127.0.0.1",
                f"--user-data-dir={TOPAS_DEBUG_PROFILE_DIR}",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-session-crashed-bubble",
                TOPAS_DEBUG_URL,
            ]
            subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            self._log(f"TOPAS 디버깅 브라우저 열기: {TOPAS_DEBUG_URL} ({TOPAS_DEBUG_ADDRESS})")
        except Exception as exc:
            messagebox.showerror("브라우저 켜기 실패", str(exc))
            self._log(f"브라우저 켜기 실패: {exc}")

    def choose_topas_debug_browser_address(self) -> str | None:
        running_without_topas: list[str] = []
        checked_errors: list[str] = []
        for address in TOPAS_DEBUG_FALLBACK_ADDRESSES:
            status = topas_debug_browser_status(address, timeout=0.6)
            if status["running"] and status["topasTargets"]:
                self._log(f"기존 TOPAS 디버깅 브라우저 재사용: {address}")
                return address
            if status["running"]:
                running_without_topas.append(address)
            else:
                checked_errors.append(f"{address}: {status.get('error')}")

        if running_without_topas:
            messagebox.showerror(
                "TOPAS 브라우저 확인",
                "디버깅 Chrome은 켜져 있지만 TOPAS 탭을 찾지 못했습니다.\n"
                f"확인한 주소: {', '.join(running_without_topas)}\n\n"
                "[브라우저 켜기]로 TOPAS 탭을 열거나, 디버깅 Chrome 안에서 TOPAS에 로그인한 뒤 다시 실행해 주세요.",
            )
        else:
            detail = "\n".join(checked_errors)
            messagebox.showerror(
                "TOPAS 브라우저 확인",
                "실행은 새 브라우저를 자동으로 띄우지 않고 기존 디버깅 Chrome을 재사용합니다.\n"
                "[브라우저 켜기]를 한 번 눌러 TOPAS에 로그인한 뒤 다시 실행해 주세요."
                + (f"\n\n확인 결과:\n{detail}" if detail else ""),
            )
        return None

    def collect_topas(self, direction: str) -> None:
        try:
            log_dir = self.require_log_dir()
            limit = AUTO_COLLECT_BATCH_SIZE
            timeout = float(self.collect_timeout.get())
            command_filter = self.automatic_collect_filter(direction)
            if timeout < 10:
                raise ValueError("timeout은 10초 이상을 권장합니다.")
            run_doc = json.loads((log_dir / "run.json").read_text(encoding="utf-8"))
            mismatches = self.run_input_mismatches(run_doc)
            if mismatches:
                raise ValueError(
                    "현재 화면 조건과 선택된 세이브가 다릅니다.\n"
                    + "\n".join(mismatches)
                    + "\n\n새 날짜/노선으로 시작하려면 '새 실행'을 누르고, 기존 작업을 이어가려면 '작업 불러오기' 후 '이어 실행'을 눌러 주세요."
                )
            debugger_address = self.choose_topas_debug_browser_address()
            if debugger_address is None:
                return
        except Exception as exc:
            messagebox.showerror("수집 조건 확인", str(exc))
            return

        title_map = {
            "all": "TOPAS 이어 실행",
            "departure": "출발 TOPAS 수집",
            "return": "귀국 TOPAS 수집",
        }
        title = title_map.get(direction, "TOPAS 수집")
        command = [
            *collector_command_prefix(),
            str(log_dir),
            "--direction",
            direction,
            "--limit",
            str(limit),
            "--filter",
            command_filter,
            "--skip-existing",
            "--timeout",
            str(timeout),
            "--live",
            "--process-after",
            "--calculate-after",
            "--loop-until-done",
            "--retry-collect-errors",
            "--max-collect-error-retries",
            str(DEFAULT_MAX_COLLECT_ERROR_RETRIES),
            "--batch-pause",
            "1",
        ]
        command.extend(["--debugger-address", debugger_address])
        self.run_background_command(title, command, controllable=True)

    def automatic_collect_filter(self, direction: str) -> str:
        return "raw-missing"

    def auto_run_all(self) -> None:
        result = self.create_run_from_current_inputs("실행 준비 실패", write_initial_excel=False)
        if result is None:
            return
        output_path, run_log = result
        self.current_log_dir.set(str(run_log["logDir"]))
        self.update_primary_run_button()
        self._log(f"실행 세이브 생성 완료: {run_log['logDir']}")
        self._log(f"최종 엑셀 예정 경로: {output_path}")
        self._log("새 작업 전체 수집을 시작합니다.")
        self.collect_topas("all")

    def process_current_run(self) -> None:
        try:
            log_dir = self.require_log_dir()
        except Exception as exc:
            messagebox.showerror("실행 로그 확인", str(exc))
            return
        command = [
            *app_command_prefix(),
            "--process-run",
            str(log_dir),
        ]
        self.run_background_command("raw 후처리", command)

    def calculate_current_run(self) -> None:
        try:
            log_dir = self.require_log_dir()
        except Exception as exc:
            messagebox.showerror("실행 로그 확인", str(exc))
            return
        command = [
            *app_command_prefix(),
            "--calculate-fares",
            str(log_dir),
        ]
        self.run_background_command("운임 계산", command)

    def open_calculated_excel(self) -> None:
        try:
            run_doc = self.read_current_run_doc()
            path_text = str(run_doc.get("calculatedExcel") or "")
            if not path_text:
                raise FileNotFoundError("최종 계산 엑셀이 아직 생성되지 않았습니다. 모든 raw 수집이 끝난 뒤 자동 생성됩니다.")
            path = Path(path_text)
            if not path.exists():
                raise FileNotFoundError(f"계산 엑셀을 찾지 못했습니다: {path}")
            os.startfile(path)
            self._log(f"계산 엑셀 열기: {path}")
        except Exception as exc:
            messagebox.showerror("계산 엑셀 열기 실패", str(exc))
            self._log(f"계산 엑셀 열기 실패: {exc}")

    def require_log_dir(self) -> Path:
        value = self.current_log_dir.get().strip()
        if not value:
            raise FileNotFoundError("작업을 먼저 실행하거나 '작업 불러오기'로 기존 작업을 불러와 주세요.")
        path = Path(value)
        if not path.exists() or not path.is_dir():
            raise FileNotFoundError(f"작업 폴더를 찾지 못했습니다: {path}")
        if not (path / "run.json").exists():
            raise FileNotFoundError(f"run.json이 없는 폴더입니다: {path}")
        return path

    def read_current_run_doc(self) -> dict[str, object]:
        log_dir = self.require_log_dir()
        return json.loads((log_dir / "run.json").read_text(encoding="utf-8"))

    def load_run_inputs(self, log_dir: Path) -> None:
        try:
            run_doc = json.loads((log_dir / "run.json").read_text(encoding="utf-8"))
        except Exception as exc:
            self._log(f"세이브 조건 불러오기 실패: {exc}")
            return
        self.start_date.set(str(run_doc.get("startDate") or self.start_date.get()))
        self.end_date.set(str(run_doc.get("endDate") or self.end_date.get()))
        self.product_days.set(str(run_doc.get("productDays") or self.product_days.get()))
        selected_routes = {str(route) for route in run_doc.get("selectedRoutes", []) if route}
        if selected_routes:
            for route, var in self.route_vars.items():
                var.set(route in selected_routes)
            selected_master_keys = {str(key) for key in run_doc.get("selectedMasterKeys", []) if key}
            for master in self.masters:
                var = self.master_vars.get(master.key)
                if var is None:
                    continue
                if selected_master_keys:
                    var.set(master.key in selected_master_keys)
                else:
                    var.set(master.route in selected_routes)
            self.update_master_detail_states()
            self._refresh_route_summary()
        self._log("세이브 조건을 화면에 반영했습니다.")

    def run_input_mismatches(self, run_doc: dict[str, object]) -> list[str]:
        mismatches: list[str] = []
        screen_start = normalize_date_input(self.start_date.get())
        screen_end = normalize_date_input(self.end_date.get())
        self.start_date.set(screen_start)
        self.end_date.set(screen_end)
        if str(run_doc.get("startDate", "")) != screen_start:
            mismatches.append(f"- 시작일: 화면 {screen_start} / 세이브 {run_doc.get('startDate', '')}")
        if str(run_doc.get("endDate", "")) != screen_end:
            mismatches.append(f"- 종료일: 화면 {screen_end} / 세이브 {run_doc.get('endDate', '')}")
        if str(run_doc.get("productDays", "")) != self.product_days.get().strip():
            mismatches.append(f"- 상품일수: 화면 {self.product_days.get().strip()} / 세이브 {run_doc.get('productDays', '')}")
        run_routes = set(str(route) for route in run_doc.get("selectedRoutes", []) if route)
        screen_masters = self.selected_masters()
        screen_routes = {master.route for master in screen_masters}
        if run_routes != screen_routes:
            mismatches.append(
                "- 노선: 화면 "
                + ", ".join(sorted(screen_routes))
                + " / 세이브 "
                + ", ".join(sorted(run_routes))
            )
        run_master_keys = {str(key) for key in run_doc.get("selectedMasterKeys", []) if key}
        if run_master_keys:
            screen_master_keys = {master.key for master in screen_masters}
            if run_master_keys != screen_master_keys:
                mismatches.append(
                    "- 항공사/편명: 화면 "
                    + ", ".join(sorted(screen_master_keys))
                    + " / 세이브 "
                    + ", ".join(sorted(run_master_keys))
                )
        return mismatches

    def set_action_buttons_state(self, state: str) -> None:
        for button in self.action_buttons:
            button.configure(state=state)

    def control_file(self, filename: str) -> Path:
        return self.require_log_dir() / filename

    def clear_collection_control_files(self) -> None:
        value = self.current_log_dir.get().strip()
        if not value:
            return
        log_dir = Path(value)
        for filename in (PAUSE_CONTROL_FILE, STOP_CONTROL_FILE):
            path = log_dir / filename
            try:
                if path.exists():
                    path.unlink()
            except Exception as exc:
                self._log(f"제어 파일 정리 실패: {path} | {exc}")

    def set_collection_control_state(self, running: bool) -> None:
        self._background_running = running
        if running or not self._command_running:
            self.update_primary_run_button()
        if self.stop_button is not None:
            self.stop_button.configure(text="■ 정지", state=tk.NORMAL if running else tk.DISABLED)

    def toggle_pause_collection(self) -> None:
        if not self._background_running:
            return
        try:
            pause_path = self.control_file(PAUSE_CONTROL_FILE)
            stop_path = self.control_file(STOP_CONTROL_FILE)
            if stop_path.exists():
                self._log("정지 요청 처리 중입니다.")
                return
            if pause_path.exists():
                pause_path.unlink()
                self.update_primary_run_button()
                self._log("재개 요청: 다음 묶음부터 수집을 이어갑니다.")
            else:
                pause_path.write_text(now_iso() + "\n", encoding="utf-8")
                self.update_primary_run_button()
                self._log("일시중지 요청: 현재 TOPAS 묶음 저장 후 대기합니다.")
        except Exception as exc:
            messagebox.showerror("일시중지 처리 실패", str(exc))
            self._log(f"일시중지 처리 실패: {exc}")

    def stop_collection(self) -> None:
        if not self._background_running:
            return
        try:
            stop_path = self.control_file(STOP_CONTROL_FILE)
            pause_path = self.control_file(PAUSE_CONTROL_FILE)
            stop_path.write_text(now_iso() + "\n", encoding="utf-8")
            if pause_path.exists():
                pause_path.unlink()
            if self.primary_run_button is not None:
                self.primary_run_button.configure(text="■ 정지 중", state=tk.DISABLED)
            if self.stop_button is not None:
                self.stop_button.configure(state=tk.DISABLED)
            self._log("정지 요청: 현재 TOPAS 묶음 저장 후 종료합니다.")
        except Exception as exc:
            messagebox.showerror("정지 처리 실패", str(exc))
            self._log(f"정지 처리 실패: {exc}")

    def progress_log_dir(self) -> Path | None:
        value = self.current_log_dir.get().strip()
        if not value:
            return None
        path = Path(value)
        return path if path.exists() and path.is_dir() else None

    def read_progress_events(self, log_dir: Path) -> dict[str, object]:
        events_path = log_dir / "events.jsonl"
        summary: dict[str, object] = {
            "batch": 0,
            "collected": 0,
            "failed": 0,
            "sessionErrors": 0,
            "ac1Active": False,
            "ac1Phase": "",
            "ac1Seen": 0,
            "ac1Expected": 0,
            "ac1InFlight": 0,
            "ac1CommandCount": 0,
        }
        if not events_path.exists():
            return summary
        try:
            with events_path.open("rb") as handle:
                size = handle.seek(0, os.SEEK_END)
                handle.seek(max(0, size - 262144))
                chunk = handle.read()
            text = chunk.decode("utf-8", errors="replace")
            lines = text.splitlines()
            if size > 262144 and lines:
                lines = lines[1:]
            lines = lines[-500:]
        except Exception:
            return summary
        latest_ac1_timestamp = ""
        for line in lines:
            try:
                event = json.loads(line)
            except Exception:
                continue
            timestamp = str(event.get("timestamp") or "")
            if self._progress_start_iso and timestamp and timestamp < self._progress_start_iso:
                continue
            event_type = str(event.get("event") or "")
            if "batchIndex" in event:
                try:
                    summary["batch"] = max(int(summary["batch"]), int(event.get("batchIndex") or 0))
                except Exception:
                    pass
            if event_type == "topas_command_collected":
                summary["collected"] = int(summary["collected"]) + 1
            elif event_type == "topas_command_failed":
                summary["failed"] = int(summary["failed"]) + 1
            elif event_type == "topas_session_error":
                summary["sessionErrors"] = int(summary["sessionErrors"]) + 1
            if event_type in {
                "topas_ac1_batch_started",
                "topas_ac1_batch_progress",
                "topas_ac1_batch_collected",
                "topas_ac1_batch_fallback",
            } and timestamp >= latest_ac1_timestamp:
                latest_ac1_timestamp = timestamp
                phase = str(event.get("phase") or ("fallback" if event_type == "topas_ac1_batch_fallback" else "started"))
                summary["ac1Active"] = event_type in {"topas_ac1_batch_started", "topas_ac1_batch_progress"}
                summary["ac1Phase"] = phase
                for target, source in (
                    ("ac1Seen", "seenCount"),
                    ("ac1Expected", "expectedCount"),
                    ("ac1InFlight", "inFlightCount"),
                    ("ac1CommandCount", "commandCount"),
                ):
                    try:
                        summary[target] = int(event.get(source) or 0)
                    except Exception:
                        summary[target] = 0
        return summary

    def ac1_progress_message(self, events: dict[str, object]) -> str:
        if not bool(events.get("ac1Active")):
            return ""
        phase = str(events.get("ac1Phase") or "started")
        phase_labels = {
            "started": "AC1 묶음 시작",
            "first_an_done": "대표 날짜 AN 완료",
            "ac1_input_sending": "AC1 입력 중",
            "ac1_input_sent": "AC1 입력 완료",
            "ac1_missing_retry_sent": "AC1 추가 입력",
            "waiting_response": "AC1 응답 대기",
            "response_ready": "AC1 응답 확인",
            "raw_writing": "raw 저장 중",
            "raw_written": "raw 저장 완료",
        }
        label = phase_labels.get(phase, phase)
        try:
            seen = int(events.get("ac1Seen") or 0)
            expected = int(events.get("ac1Expected") or 0)
            command_count = int(events.get("ac1CommandCount") or 0)
        except Exception:
            return label
        if expected > 0:
            return f"{label} | AC1 응답 {seen}/{expected} | 묶음 {command_count}개"
        if command_count > 0:
            return f"{label} | 묶음 {command_count}개"
        return label

    def event_log_path(self) -> Path | None:
        log_dir = self.progress_log_dir()
        if log_dir is None:
            return None
        return log_dir / "events.jsonl"

    def count_event_log_lines(self, path: Path | None) -> int:
        if path is None or not path.exists():
            return 0
        try:
            return path.stat().st_size
        except Exception:
            return 0

    def start_event_log_monitor(self) -> None:
        self.stop_event_log_monitor(flush=False)
        events_path = self.event_log_path()
        self._event_log_dir = events_path.parent if events_path is not None else None
        self._event_log_offset = self.count_event_log_lines(events_path)
        self._event_log_remainder = ""
        self._event_log_last_ac1_seen = {}
        self._event_log_direct_collected_count = 0
        self._event_log_running = True
        self.poll_event_log(schedule=True)

    def stop_event_log_monitor(self, flush: bool = True) -> None:
        if flush:
            self.poll_event_log(schedule=False)
        self._event_log_running = False
        if self._event_log_after_id is not None:
            try:
                self.root.after_cancel(self._event_log_after_id)
            except Exception:
                pass
            self._event_log_after_id = None

    def poll_event_log(self, schedule: bool = True) -> None:
        events_path = self.event_log_path()
        if events_path is not None and events_path.exists():
            try:
                current_size = events_path.stat().st_size
                if self._event_log_dir != events_path.parent:
                    self._event_log_dir = events_path.parent
                    self._event_log_offset = current_size
                    self._event_log_remainder = ""
                    self._event_log_last_ac1_seen = {}
                    self._event_log_direct_collected_count = 0
                if self._event_log_offset > current_size:
                    self._event_log_offset = 0
                    self._event_log_remainder = ""
                with events_path.open("rb") as handle:
                    handle.seek(self._event_log_offset)
                    chunk = handle.read()
                    self._event_log_offset = handle.tell()
                text = self._event_log_remainder + chunk.decode("utf-8", errors="replace")
                new_lines = text.splitlines(keepends=True)
                self._event_log_remainder = ""
                if new_lines and not new_lines[-1].endswith(("\n", "\r")):
                    self._event_log_remainder = new_lines.pop()
            except Exception:
                new_lines = []
            for line in new_lines:
                try:
                    event = json.loads(line)
                except Exception:
                    continue
                if not isinstance(event, dict):
                    continue
                message = self.event_to_log_message(event)
                if message:
                    self._log(message)
        if schedule and self._event_log_running:
            self._event_log_after_id = self.root.after(1000, self.poll_event_log)

    def ac1_event_key(self, event: dict[str, object]) -> str:
        return "|".join(
            str(event.get(key) or "")
            for key in ("startIndex", "firstCommand", "lastCommand")
        )

    def short_event_error(self, event: dict[str, object]) -> str:
        error = str(event.get("error") or "").strip()
        if not error:
            return ""
        first_line = error.splitlines()[0].strip()
        return first_line[:180]

    def event_to_log_message(self, event: dict[str, object]) -> str:
        event_type = str(event.get("event") or "")
        if event_type == "topas_live_connected":
            return f"TOPAS 연결 확인: {event.get('debuggerAddress', '')}"
        if event_type == "topas_batch_started":
            return (
                f"[batch {event.get('batchIndex', '')}] TOPAS 조회 시작: "
                f"{event.get('batchSize', 0)}개 | {event.get('direction', '')} / {event.get('commandFilter', '')}"
            )
        if event_type == "topas_batch_finished":
            return (
                f"[batch {event.get('batchIndex', '')}] 완료: "
                f"수집 {event.get('collectedCount', 0)}, 실패 {event.get('failedCount', 0)}, "
                f"세션오류 {event.get('sessionErrorCount', 0)}"
            )
        if event_type == "topas_ac1_batch_started":
            return (
                f"AC1 묶음 시작: {event.get('commandCount', 0)}개 | "
                f"{event.get('firstCommand', '')} -> {event.get('lastCommand', '')}"
            )
        if event_type == "topas_ac1_chunk_started":
            sent_count = int(event.get("sentCount") or 0)
            chunk_size = int(event.get("chunkSize") or 0)
            expected = int(event.get("expectedCount") or 0)
            return f"AC1 10개 단위 전송 시작: {sent_count + 1}-{sent_count + chunk_size}/{expected}"
        if event_type == "topas_ac1_missing_retry":
            seen = int(event.get("seenCount") or 0)
            expected = int(event.get("expectedCount") or 0)
            chunk_seen = int(event.get("chunkSeenCount") or 0)
            chunk_expected = int(event.get("chunkExpectedCount") or 0)
            missing = int(event.get("missingCount") or 0)
            if chunk_expected:
                return (
                    f"AC1 부족 응답 자동 보정: 현재 chunk {chunk_seen}/{chunk_expected} 확인, "
                    f"AC1 {missing}개 추가 전송 | 전체 {seen}/{expected}"
                )
            return f"AC1 부족 응답 자동 보정: {seen}/{expected} 확인, AC1 {missing}개 추가 전송"
        if event_type == "topas_ac1_blank_retry":
            seen = int(event.get("seenCount") or 0)
            expected = int(event.get("expectedCount") or 0)
            chunk_seen = int(event.get("chunkSeenCount") or 0)
            chunk_expected = int(event.get("chunkExpectedCount") or 0)
            retry = int(event.get("retryCount") or 0)
            target = str(event.get("expectedDate") or "")
            return (
                f"AC1 응답 없음 재전송: {target} | "
                f"현재 chunk {chunk_seen}/{chunk_expected}, 전체 {seen}/{expected}, 재시도 {retry}"
            )
        if event_type == "topas_ac1_batch_progress":
            phase = str(event.get("phase") or "")
            key = self.ac1_event_key(event)
            expected = int(event.get("expectedCount") or 0)
            seen = int(event.get("seenCount") or 0)
            command_count = int(event.get("commandCount") or 0)
            chunk_start = int(event.get("chunkStart") or 0)
            chunk_size = int(event.get("chunkSize") or 0)
            if phase == "first_an_done":
                return f"대표 날짜 AN 완료: {event.get('firstCommand', '')}"
            if phase == "ac1_input_sent":
                if chunk_start and chunk_size:
                    return f"AC1 10개 단위 입력 완료: {chunk_start}-{chunk_start + chunk_size - 1}/{expected}"
                return f"AC1 입력 완료: {expected}개"
            if phase == "waiting_response":
                last_seen = self._event_log_last_ac1_seen.get(key, -1)
                should_log = seen == 0 or seen >= expected or seen - last_seen >= 10
                if should_log:
                    self._event_log_last_ac1_seen[key] = seen
                    return f"AC1 응답 확인 중: {seen}/{expected}"
                return ""
            if phase == "response_ready":
                self._event_log_last_ac1_seen[key] = seen
                return f"AC1 응답 확인 완료: {seen}/{expected}"
            if phase == "raw_writing":
                return f"raw 저장 시작: {command_count}개"
            if phase == "raw_written":
                return f"raw 저장 완료: {command_count}개"
            return ""
        if event_type == "topas_ac1_batch_collected":
            return (
                f"AC1 묶음 완료: 수집 {event.get('commandCount', 0)}개 | "
                f"raw {event.get('rawFileWriteCount', 0)}개 / {event.get('rawWriteSeconds', 0)}초 | "
                f"전체 {event.get('elapsedSeconds', 0)}초"
            )
        if event_type == "topas_ac1_batch_fallback":
            return (
                f"AC1 묶음 실패, AN 개별조회로 전환: {event.get('commandCount', 0)}개 | "
                f"{self.short_event_error(event)}"
            )
        if event_type == "topas_command_collected":
            if not str(event.get("collectionMode") or "").startswith("direct"):
                return ""
            self._event_log_direct_collected_count += 1
            count = self._event_log_direct_collected_count
            if count == 1 or count % 10 == 0:
                return f"AN 개별조회 진행: {count}건 완료 | {event.get('command', '')}"
            return ""
        if event_type == "topas_command_failed":
            return f"TOPAS 명령 실패: {event.get('command', '')} | {self.short_event_error(event)}"
        if event_type == "topas_session_error":
            phase = str(event.get("phase") or "")
            prefix = f"TOPAS 세션 오류({phase})" if phase else "TOPAS 세션 오류"
            return f"{prefix}: {self.short_event_error(event)}"
        if event_type == "topas_collection_paused":
            return "일시중지 상태: 현재 묶음 완료 후 다음 묶음 시작 전 대기합니다."
        if event_type == "topas_collection_resumed":
            return "재개 요청 확인: TOPAS 수집을 이어갑니다."
        if event_type == "topas_collection_stop_requested":
            return f"정지 요청 감지: {event.get('phase', '')}"
        if event_type == "topas_collection_stopped":
            return f"TOPAS 수집 중단: {event.get('reason', '')} {self.short_event_error(event)}".strip()
        if event_type == "topas_session_error_circuit_opened":
            return f"세션 오류 반복으로 중단: {self.short_event_error(event)}"
        if event_type == "topas_collection_no_remaining":
            return f"TOPAS 수집 완료: 남은 조회 없음 | 총 batch {event.get('totalBatches', '')}"
        if event_type == "topas_final_pending_retry_started":
            pending = event.get("pendingSummary", {})
            pending_count = pending.get("actionablePendingCommands", 0) if isinstance(pending, dict) else 0
            return f"최종 엑셀 전 남은 조회 재시도 시작: {event.get('commandCount', pending_count)}건"
        if event_type == "topas_final_pending_retry_finished":
            pending = event.get("pendingSummary", {})
            pending_count = pending.get("actionablePendingCommands", 0) if isinstance(pending, dict) else 0
            return (
                f"남은 조회 재시도 완료: 수집 {event.get('collectedCount', 0)}, "
                f"실패 {event.get('failedCount', 0)}, 세션오류 {event.get('sessionErrorCount', 0)}, "
                f"남은 pending {pending_count}"
            )
        if event_type == "topas_final_pending_retry_failed":
            return f"남은 조회 재시도 실패: {self.short_event_error(event)}"
        if event_type == "topas_final_processing_started":
            if event.get("forcedExcelWithPending"):
                pending = event.get("pendingSummary", {})
                pending_count = pending.get("actionablePendingCommands", 0) if isinstance(pending, dict) else 0
                return f"미수집/조회오류 {pending_count}건 표시 후 최종 엑셀 생성 시작"
            return "최종 후처리/운임 계산 시작"
        if event_type == "raw_outputs_processed":
            return (
                f"raw 후처리 완료: 출발 {event.get('departureResultCount', 0)}, "
                f"귀국후보 {event.get('returnCandidateResultCount', 0)}, "
                f"박수확정 {event.get('returnNightResultCount', 0)}"
            )
        if event_type == "fare_results_calculated":
            message = (
                f"운임 계산 완료: 전체 {event.get('fareRowCount', 0)}, "
                f"가격확정 {event.get('pricedCount', 0)}, 마감 {event.get('closedCount', 0)}, "
                f"pending {event.get('actionablePendingCommands', 0)}"
            )
            if event.get("forcedExcelWithPending"):
                message += " | pending 포함 생성"
            if event.get("excelGenerated"):
                message += f" | 최종 엑셀 {event.get('calculatedExcel', '')}"
            return message
        if event_type == "topas_final_processing_finished":
            if event.get("forcedExcelWithPending"):
                return "미수집/조회오류 표시 포함 최종 후처리/운임 계산 완료"
            return "최종 후처리/운임 계산 완료"
        if event_type == "run_failed":
            return f"실행 실패: {self.short_event_error(event)}"
        return ""

    def refresh_progress_panel(self) -> None:
        log_dir = self.progress_log_dir()
        if log_dir is None:
            self.progress_percent.set(0.0)
            self.progress_text.set("진행 대기")
            self.progress_detail.set("실행을 시작하면 진행률과 예상 남은 시간이 표시됩니다.")
            return
        try:
            pending = collect_pending_summary(log_dir)
        except Exception as exc:
            self.progress_percent.set(0.0)
            self.progress_text.set("진행 상태 확인 필요")
            self.progress_detail.set(str(exc))
            return
        total_commands = int(pending.get("uniqueTopasCommands") or 0)
        actionable = int(pending.get("actionablePendingCommands") or 0)
        completed = max(0, total_commands - actionable)
        percent = (completed / total_commands * 100) if total_commands else 0.0
        self.progress_percent.set(min(100.0, max(0.0, percent)))
        self.progress_text.set(
            f"현재 세이브 진행률 {percent:.1f}% | 완료/비대상 {completed} / 전체 {total_commands} | 자동 진행대상 {actionable}"
        )
        self.progress_detail.set(
            "정상 raw {normal_raw}, 비운항 {no_flight}, 미수집 {missing}, 재시도대상 {retryable}, retry초과 {retry_exhausted}".format(
                retryable=int(pending.get("collect_error_retryable") or 0)
                + int(pending.get("empty_or_truncated_retryable") or 0)
                + int(pending.get("command_mismatch_retryable") or 0),
                **pending,
            )
        )

    def start_progress_monitor(self, title: str) -> None:
        self._progress_running = True
        self._progress_title = title
        self._progress_start_monotonic = time.monotonic()
        self._progress_start_iso = now_iso()
        self._progress_after_id = None
        self._progress_initial_pending = None
        log_dir = self.progress_log_dir()
        if log_dir is not None:
            try:
                pending = collect_pending_summary(log_dir)
                initial_pending = int(pending.get("actionablePendingCommands") or 0)
                self._progress_initial_pending = initial_pending if initial_pending > 0 else None
            except Exception:
                self._progress_initial_pending = None
        self.update_live_progress()

    def update_live_progress(self) -> None:
        if not self._progress_running:
            return
        log_dir = self.progress_log_dir()
        elapsed = (
            time.monotonic() - self._progress_start_monotonic
            if self._progress_start_monotonic is not None
            else 0.0
        )
        if log_dir is None:
            self.progress_percent.set(0.0)
            self.progress_text.set(f"{self._progress_title} 실행 중 | 경과 {format_duration(elapsed)}")
            self.progress_detail.set("선택된 실행 로그가 없어 상세 진행률을 계산하지 못했습니다.")
        else:
            try:
                pending = collect_pending_summary(log_dir)
                events = self.read_progress_events(log_dir)
                current_pending = int(pending.get("actionablePendingCommands") or 0)
                total = self._progress_initial_pending
                if total and total > 0:
                    done = max(0, min(total, total - current_pending))
                    ac1_in_flight = 0
                    if bool(events.get("ac1Active")):
                        try:
                            ac1_in_flight = max(0, int(events.get("ac1InFlight") or 0))
                        except Exception:
                            ac1_in_flight = 0
                    display_done = min(total, done + min(current_pending, ac1_in_flight))
                    percent = display_done / total * 100
                    eta_base_done = display_done if display_done > 0 else done
                    eta_seconds = (elapsed / eta_base_done * max(0, total - display_done)) if eta_base_done > 0 and current_pending > 0 else None
                    if current_pending == 0:
                        eta_text = "수집 완료, 후처리/엑셀 생성 중"
                    else:
                        eta_text = f"예상 남은 시간 {format_duration(eta_seconds)}"
                    self.progress_percent.set(min(100.0, max(0.0, percent)))
                    if display_done > done:
                        self.progress_text.set(
                            f"{self._progress_title} {percent:.1f}% | 완료 {done} + 진행중 {display_done - done} / 시작대상 {total} | 남은 {current_pending} | {eta_text}"
                        )
                    else:
                        self.progress_text.set(
                            f"{self._progress_title} {percent:.1f}% | 완료 {done} / 시작대상 {total} | 남은 {current_pending} | {eta_text}"
                        )
                else:
                    self.progress_percent.set(0.0)
                    self.progress_text.set(
                        f"{self._progress_title} 실행 중 | 자동 진행대상 {current_pending} | 경과 {format_duration(elapsed)}"
                    )
                retryable = (
                    int(pending.get("collect_error_retryable") or 0)
                    + int(pending.get("empty_or_truncated_retryable") or 0)
                    + int(pending.get("command_mismatch_retryable") or 0)
                )
                ac1_message = self.ac1_progress_message(events)
                ac1_prefix = f"{ac1_message} | " if ac1_message else ""
                self.progress_detail.set(
                    ac1_prefix
                    + "batch {batch} | 이번 실행 수집 {collected}, 실패 {failed}, 세션오류 {sessionErrors} | "
                    "정상 raw {normal_raw}, 비운항 {no_flight}, 미수집 {missing}, 재시도대상 {retryable}, retry초과 {retry_exhausted}".format(
                        retryable=retryable,
                        **events,
                        **pending,
                    )
                )
            except Exception as exc:
                self.progress_text.set(f"{self._progress_title} 실행 중 | 경과 {format_duration(elapsed)}")
                self.progress_detail.set(f"진행률 계산 실패: {exc}")
        self._progress_after_id = self.root.after(2000, self.update_live_progress)

    def finish_progress_monitor(self, title: str, returncode: int) -> None:
        self._progress_running = False
        if self._progress_after_id is not None:
            try:
                self.root.after_cancel(self._progress_after_id)
            except Exception:
                pass
            self._progress_after_id = None
        self.refresh_progress_panel()
        if returncode == 0:
            if self.progress_percent.get() >= 99.9:
                self.progress_text.set(f"{title} 완료 | 진행률 100.0%")
            else:
                self.progress_text.set(f"{title} 완료 | {self.progress_text.get()}")
            return
        self.progress_text.set(f"{title} 실패 | exit {returncode}")

    def run_background_command(self, title: str, command: list[str], controllable: bool = False) -> None:
        self.set_action_buttons_state(tk.DISABLED)
        if controllable:
            self.clear_collection_control_files()
        self._command_running = True
        self._close_after_background = False
        self.set_collection_control_state(controllable)
        self.start_progress_monitor(title)
        self.start_event_log_monitor()
        self._log(f"{title} 시작")

        def worker() -> None:
            env = os.environ.copy()
            env["PYTHONUTF8"] = "1"
            if getattr(sys, "frozen", False) and command:
                try:
                    if Path(command[0]).resolve() == Path(sys.executable).resolve():
                        env["PYINSTALLER_RESET_ENVIRONMENT"] = "1"
                except Exception:
                    env["PYINSTALLER_RESET_ENVIRONMENT"] = "1"
            process = subprocess.Popen(
                command,
                cwd=str(BASE_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
            )
            self._background_process = process
            stdout, stderr = process.communicate()
            completed = subprocess.CompletedProcess(command, process.returncode, stdout, stderr)
            self._background_process = None
            self.root.after(0, lambda: self.finish_background_command(title, completed))

        threading.Thread(target=worker, daemon=True).start()

    def finish_background_command(self, title: str, completed: subprocess.CompletedProcess[str]) -> None:
        should_close = self._close_after_background
        self._command_running = False
        self._background_process = None
        self._close_after_background = False
        self.set_action_buttons_state(tk.NORMAL)
        self.clear_collection_control_files()
        self.set_collection_control_state(False)
        self.stop_event_log_monitor(flush=True)
        self.finish_progress_monitor(title, completed.returncode)
        stdout = (completed.stdout or "").strip()
        stderr = (completed.stderr or "").strip()
        if stdout and (completed.returncode != 0 or "TOPAS" not in title):
            self._log(stdout[-3000:])
        if stderr:
            self._log(stderr[-3000:])
        if completed.returncode == 0:
            self._log(f"{title} 완료")
            self.log_current_run_summary()
            if should_close:
                self.root.after(300, self.root.destroy)
            return
        self._log(f"{title} 실패: exit {completed.returncode}")
        messagebox.showerror(title, stderr or stdout or f"exit {completed.returncode}")
        if should_close:
            self.root.after(300, self.root.destroy)

    def log_current_run_summary(self) -> None:
        value = self.current_log_dir.get().strip()
        if not value:
            return
        log_dir = Path(value)
        self.log_run_summary(log_dir, target="auto")

    def log_run_summary(self, log_dir: Path, target: str = "auto") -> None:
        if not log_dir.exists():
            return
        summaries: list[str] = []
        try:
            pending = collect_pending_summary(log_dir)
            summaries.append(
                "진행: 논리 후보 {logicalRows}, 실제 TOPAS {uniqueTopasCommands}, unique raw {uniqueRawFiles}, 자동 진행대상 {actionablePendingCommands}".format(
                    **pending
                )
            )
        except Exception:
            pass
        for filename, label in (
            ("departure-results.json", "출발"),
            ("return-night-results.json", "귀국박수"),
            ("fare-results.json", "운임"),
        ):
            path = log_dir / filename
            if not path.exists():
                continue
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            rows = payload.get("rows", []) if isinstance(payload, dict) else []
            if not isinstance(rows, list):
                continue
            counts: dict[str, int] = {}
            for row in rows:
                if not isinstance(row, dict):
                    continue
                status = fare_status_label(str(row.get("status") or "unknown"))
                counts[status] = counts.get(status, 0) + 1
            if counts:
                summary = ", ".join(f"{status} {count}" for status, count in sorted(counts.items()))
                summaries.append(f"{label}: {summary}")
        if summaries:
            self._log_to_target("현재 run 요약 | " + " | ".join(summaries), target)

    def on_close(self) -> None:
        if not self._command_running:
            self.root.destroy()
            return
        if self._background_running:
            should_stop = messagebox.askyesno(
                "작업 실행 중",
                "TOPAS 수집이 실행 중입니다.\n\n"
                "바로 창을 닫으면 수집 프로세스가 남아 PyInstaller 임시 폴더 정리 경고가 뜰 수 있습니다.\n"
                "정지 요청 후 작업이 끝나면 창을 닫을까요?",
            )
            if should_stop:
                self.stop_collection()
                self._close_after_background = True
                self._log("정지 요청 후 작업이 끝나면 창을 닫습니다.")
            return
        messagebox.showinfo(
            "작업 실행 중",
            "백그라운드 작업이 실행 중입니다.\n작업이 완료된 뒤 창을 닫아 주세요.",
        )

    def create_run_from_current_inputs(
        self,
        error_title: str,
        write_initial_excel: bool = True,
    ) -> tuple[Path, dict[str, object]] | None:
        run_log: dict[str, object] | None = None
        try:
            start = parse_iso_date(self.start_date.get())
            end = parse_iso_date(self.end_date.get())
            self.start_date.set(start.isoformat())
            self.end_date.set(end.isoformat())
            product_days = int(self.product_days.get())
            if end < start:
                raise ValueError("조회 종료일은 조회 시작일보다 빠를 수 없습니다.")
            count = date_count(start, end)
            if not (1 <= count <= 366):
                raise ValueError("조회 날짜 수는 1~366일 사이로 입력해 주세요.")
            if not (2 <= product_days <= 30):
                raise ValueError("상품일수는 2~30 사이로 입력해 주세요.")
            masters = self.selected_masters()
            if not masters:
                raise ValueError("조회할 노선과 항공사/편명을 1개 이상 선택해 주세요.")

            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            output_path = OUTPUT_DIR / default_output_path().name
            run_log = create_run_log(start, end, product_days, masters, output_path)
            command_plan = build_command_plan(str(run_log["runId"]), masters, start, end, product_days)
            write_command_plan(run_log, command_plan)
            if write_initial_excel:
                append_run_event(
                    run_log,
                    "excel_generation_started",
                    routeCount=len(run_log["selectedRoutes"]),
                    flightMasterCount=len(masters),
                )
                output_path = write_route_mvp_excel(output_path, masters, start, end)
                finish_run_log(
                    run_log,
                    "completed",
                    outputExcel=str(output_path),
                    excelGenerated=True,
                )
            else:
                append_run_event(
                    run_log,
                    "run_prepared_for_collection",
                    routeCount=len(run_log["selectedRoutes"]),
                    flightMasterCount=len(masters),
                    plannedExcel=str(output_path),
                )
                run_log.update(
                    status="prepared",
                    preparedAt=now_iso(),
                    outputExcel=str(output_path),
                    excelGenerated=False,
                )
                write_run_json(run_log)
            return output_path, run_log
        except Exception as exc:
            if run_log is not None:
                finish_run_log(run_log, "failed", error=str(exc))
            messagebox.showerror(error_title, str(exc))
            self._log(f"실패: {exc}")
            return None

    def generate_excel(self) -> None:
        result = self.create_run_from_current_inputs("엑셀 생성 실패")
        if result is None:
            return
        output_path, run_log = result

        self._log(f"엑셀 생성 완료: {output_path}")
        self._log(f"실행 로그: {run_log['logDir']}")
        self.current_log_dir.set(str(run_log["logDir"]))
        self.update_primary_run_button()
        messagebox.showinfo("엑셀 생성 완료", f"생성 완료:\n{output_path}")

    def _write_log(self, widget: tk.Text | None, message: str) -> None:
        if widget is None:
            return
        widget.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        widget.see(tk.END)

    def _log(self, message: str) -> None:
        self._write_log(self.log, message)

    def _log_recalc(self, message: str) -> None:
        self._write_log(self.recalc_log, message)

    def _log_manifest(self, message: str) -> None:
        self._write_log(self.manifest_log, message)

    def _log_to_target(self, message: str, target: str = "auto") -> None:
        if target == "recalc":
            self._log_recalc(message)
            return
        if target == "manifest":
            self._log_manifest(message)
            return
        self._log(message)


def run_sample(args: argparse.Namespace) -> Path:
    masters = [master for master in load_flight_masters() if master.enabled]
    start = parse_iso_date(args.start_date) if args.start_date else date.today()
    if args.end_date:
        end = parse_iso_date(args.end_date)
    elif args.days:
        end = start + timedelta(days=args.days - 1)
    else:
        end = default_end_date(start)
    if end < start:
        raise ValueError("조회 종료일은 조회 시작일보다 빠를 수 없습니다.")
    count = date_count(start, end)
    if not (1 <= count <= 366):
        raise ValueError("조회 날짜 수는 1~366일 사이로 입력해 주세요.")
    output_path = Path(args.output) if args.output else default_output_path()
    fare_cells = load_fare_results(args.fare_results)
    run_log = create_run_log(start, end, args.product_days, masters, output_path)
    command_plan = build_command_plan(str(run_log["runId"]), masters, start, end, args.product_days)
    write_command_plan(run_log, command_plan)
    if fare_cells:
        append_run_event(
            run_log,
            "fare_results_loaded",
            fareResultFile=str(Path(args.fare_results)),
            fareCellCount=len(fare_cells),
        )
    append_run_event(
        run_log,
        "excel_generation_started",
        routeCount=len(run_log["selectedRoutes"]),
        flightMasterCount=len(masters),
    )
    try:
        result = write_route_mvp_excel(output_path, masters, start, end, fare_cells)
    except Exception as exc:
        finish_run_log(run_log, "failed", error=str(exc))
        raise
    finish_run_log(
        run_log,
        "completed",
        outputExcel=str(result),
        excelGenerated=True,
    )
    return result


def run_embedded_collector(argv: list[str]) -> int:
    import topas_live_collector

    return topas_live_collector.main(argv)


def main() -> int:
    if len(sys.argv) > 1 and sys.argv[1] == "--collector":
        return run_embedded_collector(sys.argv[2:])

    parser = argparse.ArgumentParser(description="항공자동조회 MVP GUI / Excel generator")
    parser.add_argument("--sample", action="store_true", help="GUI 없이 샘플 엑셀 생성")
    parser.add_argument("--process-run", default="", help="기존 output/logs/{runId} raw 파일을 결과 JSON으로 후처리")
    parser.add_argument("--calculate-fares", default="", help="기존 output/logs/{runId} raw 결과와 운임 DB로 fare-results.json 계산")
    parser.add_argument("--recalculate-raw-store", default="", help="raw-store.sqlite를 재사용해 상품일수만 바꾼 계산 엑셀 생성")
    parser.add_argument("--start-date", default=date.today().isoformat(), help="조회 시작일 YYYY-MM-DD 또는 YYYYMMDD")
    parser.add_argument("--end-date", default="", help="조회 종료일 YYYY-MM-DD 또는 YYYYMMDD")
    parser.add_argument("--days", type=int, default=0, help="호환용: 종료일이 없을 때 시작일부터 N일")
    parser.add_argument("--product-days", type=int, default=5)
    parser.add_argument("--fallback-mode", choices=["before_midnight", "after_midnight", "manual"], default="before_midnight")
    parser.add_argument("--output", default="")
    parser.add_argument("--fare-results", default="", help="선택: 결과 엑셀에 반영할 요금/마감 JSON")
    parser.add_argument("--fare-cache", default="", help="선택: 운임 DB 캐시 파일 경로")
    parser.add_argument("--refresh-fare-db", action="store_true", help="운임 DB를 캐시 우선 없이 새로 로드")
    parser.add_argument("--force-excel-with-pending", action="store_true", help="미수집/조회오류가 남아도 해당 항목을 표시하고 최종 엑셀 생성")
    args = parser.parse_args()

    if args.calculate_fares:
        path = calculate_run_fares(
            args.calculate_fares,
            cache_path=args.fare_cache or None,
            refresh_fare_db=args.refresh_fare_db,
            force_excel_with_pending=args.force_excel_with_pending,
        )
        print(path)
        run_doc = json.loads((Path(args.calculate_fares) / "run.json").read_text(encoding="utf-8"))
        print(run_doc.get("calculatedExcel", ""))
        return 0

    if args.recalculate_raw_store:
        path, run_doc = recalculate_raw_store_excel(
            args.recalculate_raw_store,
            args.product_days,
            force_excel_with_pending=True,
        )
        print(path)
        print(run_doc.get("logDir", ""))
        return 0

    if args.process_run:
        outputs = process_run_raw_outputs(args.process_run)
        for path in outputs.values():
            print(path)
        return 0

    if args.sample:
        path = run_sample(args)
        print(path)
        return 0

    root = tk.Tk()
    AirAutoLookupApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
